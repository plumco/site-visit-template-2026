import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.io as pio
import gspread
import requests
import json
import io
from google.oauth2.service_account import Credentials
from html import escape
from datetime import datetime
import streamlit.components.v1 as components
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.barcharts import VerticalBarChart

_glass_template = pio.templates["plotly_dark"]
_glass_template.layout.paper_bgcolor = "#13243D"
_glass_template.layout.plot_bgcolor  = "#13243D"
_glass_template.layout.font.color    = "#CBD5E1"
_glass_template.layout.font.family   = "Inter, sans-serif"
_glass_template.layout.xaxis.gridcolor = "rgba(255,255,255,0.08)"
_glass_template.layout.yaxis.gridcolor = "rgba(255,255,255,0.08)"
_glass_template.layout.xaxis.linecolor = "rgba(255,255,255,0.15)"
_glass_template.layout.yaxis.linecolor = "rgba(255,255,255,0.15)"
pio.templates["liquid_glass"] = _glass_template
px.defaults.template = "liquid_glass"

st.set_page_config(layout="wide", page_title="Site Visit Deep Analytics", page_icon="📊")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Sora:wght@400;600;700;800&family=Inter:wght@400;500;600&family=JetBrains+Mono:wght@400;500;600&display=swap');
html,body,[class*="css"]{font-family:'Inter',sans-serif;}
.stApp{background:radial-gradient(circle at 12% 18%,rgba(56,189,248,0.30) 0%,transparent 40%),radial-gradient(circle at 88% 6%,rgba(129,140,248,0.26) 0%,transparent 38%),radial-gradient(circle at 50% 100%,rgba(34,211,238,0.22) 0%,transparent 45%),radial-gradient(circle at 30% 70%,rgba(99,102,241,0.16) 0%,transparent 40%),linear-gradient(180deg,#060B16 0%,#0B1220 55%,#0A1020 100%) !important;background-attachment:fixed !important;}
[data-testid="stAppViewContainer"],[data-testid="stHeader"],[data-testid="stMain"],[data-testid="stToolbar"],[data-testid="stDecoration"],[data-testid="stBottomBlockContainer"]{background:transparent !important;}
h1,h2,h3,h4{font-family:'Sora',sans-serif !important;color:#F1F5F9 !important;letter-spacing:-0.01em;}
h1{font-weight:800 !important;}h2,h3{font-weight:700 !important;}
p,span,label,.stMarkdown,.stCaption,div[data-testid="stCaptionContainer"]{color:#CBD5E1;}
section[data-testid="stSidebar"]{background:rgba(15,23,42,0.55) !important;backdrop-filter:blur(24px) saturate(150%);-webkit-backdrop-filter:blur(24px) saturate(150%);border-right:1px solid rgba(255,255,255,0.08);}
section[data-testid="stSidebar"] *{color:#E2E8F0 !important;}
div[data-testid="stMetric"],div[data-testid="metric-container"]{background:linear-gradient(160deg,rgba(56,189,248,0.10),rgba(255,255,255,0.04)) !important;backdrop-filter:blur(20px) saturate(160%);-webkit-backdrop-filter:blur(20px) saturate(160%);border:1px solid rgba(56,189,248,0.20) !important;border-radius:18px !important;padding:1.4rem 1.2rem !important;box-shadow:0 8px 32px rgba(0,0,0,0.35),inset 0 1px 0 rgba(255,255,255,0.08);transition:transform 0.2s ease,box-shadow 0.2s ease,border-color 0.2s ease;}
div[data-testid="stMetric"]:hover{transform:translateY(-2px);box-shadow:0 12px 40px rgba(56,189,248,0.25),inset 0 1px 0 rgba(255,255,255,0.1);border-color:rgba(56,189,248,0.5) !important;}
div[data-testid="stMetricLabel"]{color:#94A3B8 !important;font-family:'Inter',sans-serif !important;font-size:0.78rem !important;letter-spacing:0.04em;text-transform:uppercase;}
div[data-testid="stMetricValue"]{color:#F8FAFC !important;font-family:'JetBrains Mono',monospace !important;font-weight:600 !important;}
.stTabs [data-baseweb="tab-list"]{background:rgba(255,255,255,0.04);backdrop-filter:blur(16px);border-radius:14px;padding:6px;border:1px solid rgba(255,255,255,0.08);gap:6px;}
.stTabs [data-baseweb="tab"]{color:#94A3B8 !important;background:rgba(255,255,255,0.03) !important;border:1px solid rgba(255,255,255,0.08) !important;border-radius:10px !important;padding:8px 16px !important;font-family:'Sora',sans-serif !important;font-weight:600 !important;font-size:0.9rem !important;transition:all 0.18s ease;}
.stTabs [data-baseweb="tab"]:hover{background:rgba(56,189,248,0.08) !important;border-color:rgba(56,189,248,0.25) !important;color:#CBD5E1 !important;}
.stTabs [aria-selected="true"]{background:rgba(56,189,248,0.18) !important;border-color:rgba(56,189,248,0.45) !important;color:#7DD3FC !important;box-shadow:0 0 16px rgba(56,189,248,0.15);}
.stTabs [data-baseweb="tab-highlight"],.stTabs [data-baseweb="tab-border"]{display:none !important;height:0 !important;background:transparent !important;}
.stButton button,.stDownloadButton button,.stFormSubmitButton button,button[kind="secondary"],button[kind="primary"],button[data-testid="baseButton-secondary"],button[data-testid="baseButton-primary"],button[data-testid="stBaseButton-secondary"],button[data-testid="stBaseButton-primary"]{background:rgba(56,189,248,0.14) !important;backdrop-filter:blur(12px);border:1px solid rgba(56,189,248,0.45) !important;color:#7DD3FC !important;border-radius:12px !important;font-family:'Sora',sans-serif !important;font-weight:600 !important;transition:all 0.2s ease;}
.stButton button:hover,.stDownloadButton button:hover,.stFormSubmitButton button:hover,button[kind="secondary"]:hover,button[kind="primary"]:hover,button[data-testid="baseButton-secondary"]:hover,button[data-testid="baseButton-primary"]:hover,button[data-testid="stBaseButton-secondary"]:hover,button[data-testid="stBaseButton-primary"]:hover{background:rgba(56,189,248,0.26) !important;box-shadow:0 0 20px rgba(56,189,248,0.3);border-color:rgba(56,189,248,0.7) !important;color:#E0F2FE !important;}
.stButton button p,.stDownloadButton button p,.stFormSubmitButton button p{color:inherit !important;}
div[data-testid="stExpander"]{background:rgba(255,255,255,0.04) !important;backdrop-filter:blur(16px);-webkit-backdrop-filter:blur(16px);border:1px solid rgba(255,255,255,0.1) !important;border-radius:14px !important;}
div[data-baseweb="select"]>div,.stTextInput input,.stTextArea textarea,.stDateInput input{background:rgba(255,255,255,0.05) !important;border:1px solid rgba(255,255,255,0.12) !important;border-radius:10px !important;color:#F1F5F9 !important;}
div[data-testid="stDataFrame"]{border-radius:14px;overflow:hidden;border:1px solid rgba(56,189,248,0.15);}
div[data-testid="stDataFrame"] *,.stApp{scrollbar-width:thin;scrollbar-color:rgba(56,189,248,0.45) rgba(255,255,255,0.05);}
::-webkit-scrollbar{width:10px;height:10px;}
::-webkit-scrollbar-track{background:rgba(255,255,255,0.04);border-radius:8px;}
::-webkit-scrollbar-thumb{background-color:rgba(56,189,248,0.40);border-radius:8px;border:2px solid transparent;background-clip:padding-box;}
::-webkit-scrollbar-thumb:hover{background-color:rgba(56,189,248,0.65);}
::-webkit-scrollbar-corner{background:transparent;}
div[data-testid="stVerticalBlockBorderWrapper"]{background:rgba(255,255,255,0.05) !important;backdrop-filter:blur(18px) saturate(150%);-webkit-backdrop-filter:blur(18px) saturate(150%);border:1px solid rgba(56,189,248,0.14) !important;border-radius:16px !important;padding:0.6rem !important;box-shadow:0 6px 24px rgba(0,0,0,0.3);}
.highlight-card{padding:22px;border-radius:16px;text-align:left;font-family:'Inter',sans-serif;margin-top:10px;backdrop-filter:blur(20px) saturate(160%);-webkit-backdrop-filter:blur(20px) saturate(160%);box-shadow:0 8px 28px rgba(0,0,0,0.3),inset 0 1px 0 rgba(255,255,255,0.08);border:1px solid;}
.card-blue{background:rgba(56,189,248,0.10);border-color:rgba(56,189,248,0.30);color:#7DD3FC;}
.card-green{background:rgba(52,211,153,0.10);border-color:rgba(52,211,153,0.30);color:#6EE7B7;}
.card-red{background:rgba(248,113,113,0.10);border-color:rgba(248,113,113,0.30);color:#FCA5A5;}
.card-title{font-size:0.82rem;margin-bottom:6px;opacity:0.9;font-weight:600;letter-spacing:0.02em;}
.card-value{font-size:1.15rem;font-family:'JetBrains Mono',monospace;font-weight:600;color:#F1F5F9;}
</style>""", unsafe_allow_html=True)

FIREBASE_API_KEY = "AIzaSyDuf1MozrcpQmlnbJXa-bm5C2htxRzeZOA"

def firebase_sign_in(email, password):
    url = f"https://identitytoolkit.googleapis.com/v1/accounts:signInWithPassword?key={FIREBASE_API_KEY}"
    try:
        r = requests.post(url, json={"email":email,"password":password,"returnSecureToken":True}, timeout=10)
        d = r.json()
        if r.status_code == 200:
            return True, d.get("email", email), None
        return False, None, d.get("error",{}).get("message","Authentication failed")
    except Exception as e:
        return False, None, str(e)

if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False
    st.session_state["user_email"] = None

if not st.session_state["authenticated"]:
    st.markdown("""<style>
    .login-header{text-align:center;padding:2.5rem 0 1.5rem 0;}
    .login-header h1{font-family:'Sora',sans-serif;font-size:2.4rem;font-weight:800;background:linear-gradient(135deg,#38BDF8,#818CF8);-webkit-background-clip:text;-webkit-text-fill-color:transparent;margin-bottom:0.5rem;}
    .login-header p{color:#94A3B8;font-size:1.05rem;font-family:'Inter',sans-serif;}
    div[data-testid="stForm"]{background:rgba(255,255,255,0.05) !important;backdrop-filter:blur(28px) saturate(160%);-webkit-backdrop-filter:blur(28px) saturate(160%);border:1px solid rgba(255,255,255,0.14) !important;border-radius:20px !important;padding:2.2rem !important;box-shadow:0 20px 60px rgba(0,0,0,0.4),inset 0 1px 0 rgba(255,255,255,0.1);}
    </style>""", unsafe_allow_html=True)
    st.markdown('<div class="login-header"><h1>🔐 Site Visit Analytics</h1><p>Please sign in to access the dashboard</p></div>', unsafe_allow_html=True)
    _, cc, _ = st.columns([1,1.5,1])
    with cc:
        with st.form("login_form", clear_on_submit=False):
            em = st.text_input("📧 Email", placeholder="you@example.com")
            pw = st.text_input("🔑 Password", type="password", placeholder="Your password")
            st.markdown("")
            if st.form_submit_button("🚀 Sign In", use_container_width=True):
                if not em or not pw:
                    st.error("Please enter both email and password.")
                else:
                    with st.spinner("Authenticating..."):
                        ok, ue, err = firebase_sign_in(em, pw)
                    if ok:
                        st.session_state["authenticated"] = True
                        st.session_state["user_email"] = ue
                        st.rerun()
                    else:
                        em_map = {"INVALID_LOGIN_CREDENTIALS":"❌ Invalid email or password.","EMAIL_NOT_FOUND":"❌ No account found.","INVALID_PASSWORD":"❌ Incorrect password.","USER_DISABLED":"❌ Account disabled.","TOO_MANY_ATTEMPTS_TRY_LATER":"❌ Too many attempts."}
                        st.error(em_map.get(err, f"❌ Login failed: {err}"))
    st.stop()

st.sidebar.markdown(f"👤 **{st.session_state.get('user_email','')}**")
if st.sidebar.button("🚪 Logout"):
    st.session_state["authenticated"] = False
    st.session_state["user_email"] = None
    st.cache_data.clear(); st.cache_resource.clear(); st.rerun()
st.sidebar.markdown("---")

@st.cache_resource
def init_connection():
    scopes = ["https://www.googleapis.com/auth/spreadsheets","https://www.googleapis.com/auth/drive"]
    creds = Credentials.from_service_account_info(st.secrets["gcp_service_account"], scopes=scopes)
    return gspread.authorize(creds)

client = init_connection()
SHEET_URL = "https://docs.google.com/spreadsheets/d/1J1K31wLOepJMO6DPHySUGR43GpV2sV7PqSHetO_EFjo/edit?gid=502709304#gid=502709304"

def safe_text(v):
    v = str(v).strip()
    return "-" if v.lower() in ["nan","none","null","nat",""] else escape(v)

def make_unique_headers(raw):
    seen={}; out=[]
    for h in raw:
        h=str(h).strip() or "Blank"
        seen[h]=seen.get(h,0); out.append(f"{h}_{seen[h]}" if seen[h] else h); seen[h]+=1
    return out

def clean_df(df):
    if df.empty: return df
    df=df.copy(); df.columns=[str(c).strip() for c in df.columns]
    for col in df.columns:
        df[col]=df[col].astype(str).replace({"nan":"","None":"","NaT":"","NaN":"","null":"","Null":""}).str.strip()
    return df

def safe_col(df, opts):
    if df.empty: return None
    for o in opts:
        if o in df.columns: return o
    return None

def clean_options(s):
    return sorted(s.astype(str).str.strip().replace(["nan","None","NaT","","null","Null"],pd.NA).dropna().unique().tolist())

def get_visit_status(row):
    ir=str(row.get("Is Report Visit?","")).strip().lower()
    sd=str(row.get("Report Submitted Date","")).strip()
    if ir in ["no","n","false","n/a","na"]: return "Technical (NA)"
    if sd and sd.lower() not in ["nan","none","","nat"]: return "Submitted"
    return "Pending"

def parse_floor(val):
    v=str(val).strip()
    if not v or v.lower() in ["nan","none","null","n/a","na","-"]: return 0
    try: return int(float(v))
    except: return 1

def find_master_site_col(df):
    return safe_col(df,["PROJECT","Project","PROJECT NAME","Project Name","Site Name","SITE NAME","Site"])

def find_visit_site_col(df):
    return safe_col(df,["Site Name","SITE NAME","PROJECT","Project","PROJECT NAME","Project Name"])

def filter_site(df, sc, sel):
    if df.empty or not sc: return pd.DataFrame()
    return df[df[sc].astype(str).str.strip().str.lower()==str(sel).strip().lower()].copy()

def get_row_value(row, col):
    if col and col in row.index:
        v=str(row.get(col,"")).strip()
        if v and v.lower() not in ["nan","none","null","nat"]: return v
    return "-"

def build_horizontal_table(row, columns):
    h="<table class='horizontal-info-table'><tr>"+"".join(f"<th>{safe_text(l)}</th>" for l,c in columns)+"</tr><tr>"+"".join(f"<td>{safe_text(get_row_value(row,c))}</td>" for l,c in columns)+"</tr></table>"
    return h

def df_to_html_table(df):
    if df.empty: return "<p>No data found.</p>"
    h="<table border='1' style='border-collapse:collapse;width:100%;font-size:12px;'><tr>"
    h+="".join(f"<th style='background:#f3f4f6;padding:8px;text-align:left;'>{safe_text(c)}</th>" for c in df.columns)+"</tr>"
    for _,row in df.astype(str).iterrows():
        h+="<tr>"+"".join(f"<td style='padding:8px;vertical-align:top;'>{safe_text(row[c])}</td>" for c in df.columns)+"</tr>"
    return h+"</table>"

def create_excel_compatible_report(site_name,master_df,visit_df,summary_df,last_comment):
    html=f"<html><head><meta charset='UTF-8'></head><body><h2>{safe_text(site_name)} - Site Visit Report</h2><p>Generated On: {datetime.now().strftime('%d-%m-%Y %I:%M %p')}</p><h3>Site Summary</h3>{df_to_html_table(summary_df)}<h3>Last Comment</h3><p>{safe_text(last_comment)}</p><h3>MasterProject Details</h3>{df_to_html_table(master_df)}<h3>VisitLog Details</h3>{df_to_html_table(visit_df)}</body></html>"
    return html.encode("utf-8")

def create_print_html_report(site_name,master_row,mc1,mc2,summary_df,visit_df,last_comment):
    html=f"""<html><head><meta charset="UTF-8"><title>{safe_text(site_name)} Site Report</title>
<style>body{{font-family:Arial,sans-serif;margin:30px;color:#111827;}}.header{{display:flex;justify-content:space-between;border-bottom:3px solid #111827;padding-bottom:12px;margin-bottom:18px;}}.title{{font-size:28px;font-weight:800;}}.subtitle{{color:#4b5563;font-size:13px;margin-top:5px;}}.badge{{background:#111827;color:white;padding:8px 12px;border-radius:6px;font-size:12px;height:fit-content;}}h3{{border-left:5px solid #2563eb;padding-left:10px;margin-top:24px;}}table{{width:100%;border-collapse:collapse;font-size:12px;margin-bottom:16px;}}th{{background:#f3f4f6;border:1px solid #d1d5db;padding:8px;text-align:left;}}td{{border:1px solid #d1d5db;padding:8px;vertical-align:top;}}.comment{{background:#fffbeb;border:1px solid #f59e0b;padding:12px;border-radius:8px;color:#78350f;}}</style>
</head><body><div class="header"><div><div class="title">{safe_text(site_name)}</div><div class="subtitle">Site Visit Report | Generated On {datetime.now().strftime("%d-%m-%Y %I:%M %p")}</div></div><div class="badge">Huliot Site Report</div></div>
<h3>1. Site Master Information</h3>{build_horizontal_table(master_row,mc1)}{build_horizontal_table(master_row,mc2)}
<h3>2. Visit Summary</h3>{df_to_html_table(summary_df)}<h3>3. Last Visit Comment</h3><div class="comment">{safe_text(last_comment)}</div><h3>4. VisitLog Details</h3>{df_to_html_table(visit_df)}</body></html>"""
    return html.encode("utf-8")

def create_site_card_html(selected_site,mrow,mc1,mc2,tvr,tfl,sr,pr,tna,tt,lvd,lvb,lvc,filter_label="All Data"):
    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
<style>@import url('https://fonts.googleapis.com/css2?family=Sora:wght@600;700;800&family=Inter:wght@400;500;600&family=JetBrains+Mono:wght@600&display=swap');
body{{margin:0;padding:0;font-family:'Inter',Arial,sans-serif;background:radial-gradient(circle at 10% 10%,rgba(56,189,248,0.20) 0%,transparent 45%),radial-gradient(circle at 90% 90%,rgba(129,140,248,0.16) 0%,transparent 45%),#0E1B2E;}}
.card{{background:rgba(255,255,255,0.05);backdrop-filter:blur(20px);-webkit-backdrop-filter:blur(20px);color:#F1F5F9;border:1px solid rgba(56,189,248,0.20);border-radius:18px;padding:22px;box-sizing:border-box;width:100%;}}
.hdr{{display:flex;justify-content:space-between;align-items:flex-start;border-bottom:1px solid rgba(56,189,248,0.25);padding-bottom:14px;margin-bottom:18px;}}
.title{{font-family:'Sora',sans-serif;font-size:28px;font-weight:800;background:linear-gradient(135deg,#38BDF8,#818CF8);-webkit-background-clip:text;-webkit-text-fill-color:transparent;margin-bottom:5px;}}
.sub{{font-size:13px;color:#94A3B8;}}.badge{{background:rgba(56,189,248,0.15);border:1px solid rgba(56,189,248,0.4);color:#7DD3FC;padding:8px 14px;border-radius:10px;font-size:12px;font-weight:700;white-space:nowrap;}}
.sec{{font-family:'Sora',sans-serif;font-size:17px;font-weight:700;color:#F1F5F9;margin-top:20px;margin-bottom:10px;border-left:4px solid #38BDF8;padding-left:10px;}}
.horizontal-info-table{{width:100%;border-collapse:collapse;margin-bottom:15px;font-size:13px;}}
.horizontal-info-table th{{background:rgba(255,255,255,0.06);color:#94A3B8;border:1px solid rgba(255,255,255,0.1);padding:9px;text-align:left;font-weight:700;white-space:nowrap;text-transform:uppercase;font-size:11px;letter-spacing:0.03em;}}
.horizontal-info-table td{{border:1px solid rgba(255,255,255,0.1);padding:9px;color:#F1F5F9;vertical-align:top;word-break:break-word;background:rgba(255,255,255,0.02);}}
.kpi-strip{{display:grid;grid-template-columns:repeat(6,1fr);gap:10px;margin-top:12px;margin-bottom:15px;}}
.kpi{{background:rgba(56,189,248,0.08);border:1px solid rgba(56,189,248,0.20);border-radius:12px;padding:12px;}}
.kl{{font-size:11px;color:#94A3B8;font-weight:700;margin-bottom:6px;text-transform:uppercase;letter-spacing:0.03em;}}
.kv{{font-family:'JetBrains Mono',monospace;font-size:22px;font-weight:600;color:#F8FAFC;}}
.cmt{{background:rgba(251,191,36,0.08);border:1px solid rgba(251,191,36,0.30);color:#FDE68A;padding:14px;border-radius:12px;margin-top:8px;font-size:14px;}}
@media(max-width:1200px){{.kpi-strip{{grid-template-columns:repeat(3,1fr);}}}}
@media(max-width:700px){{.kpi-strip{{grid-template-columns:repeat(1,1fr);}}}}
</style></head><body>
<div class="card">
<div class="hdr"><div><div class="title">{safe_text(selected_site)}</div><div class="sub">Site Visit Report | Live Google Sheet</div></div><div class="badge">Live Google Sheet Report</div></div>
<div style="margin-bottom:14px;"><span style="background:rgba(52,211,153,0.12);border:1px solid rgba(52,211,153,0.35);color:#6EE7B7;padding:5px 12px;border-radius:20px;font-size:12px;font-weight:700;">🔍 Viewing: {safe_text(filter_label)}</span></div>
<div class="sec">1. Site Master Information</div>
{build_horizontal_table(mrow,mc1)}{build_horizontal_table(mrow,mc2)}
<div class="sec">2. Visit Summary</div>
<div class="kpi-strip">
<div class="kpi"><div class="kl">Visit Records</div><div class="kv">{tvr}</div></div>
<div class="kpi"><div class="kl">Floor Visits</div><div class="kv">{tfl}</div></div>
<div class="kpi"><div class="kl">Submitted</div><div class="kv">{sr}</div></div>
<div class="kpi"><div class="kl">Pending</div><div class="kv">{pr}</div></div>
<div class="kpi"><div class="kl">Technical NA</div><div class="kv">{tna}</div></div>
<div class="kpi"><div class="kl">Towers</div><div class="kv">{tt}</div></div>
</div>
<div class="sec">3. Last Visit Comment</div>
<div class="cmt"><b>Date:</b> {safe_text(lvd)} &nbsp;|&nbsp; <b>Visited By:</b> {safe_text(lvb)}<br><br><b>Comment:</b> {safe_text(lvc)}</div>
</div></body></html>"""

# ==========================================
# ISSUE TRACKER CONSTANTS
# ==========================================
ISSUES_SHEET_NAME = "SiteIssues"
ISSUE_HEADERS = ["Issue ID","Site Name","Issue Type","Severity","Description","Raised By","Raised Date","Assigned To","Target Date","Status","Resolution Notes","Created At","Updated At"]
ISSUE_TYPES = ["Installation Defect","Material Issue","Design Non-compliance","Slope / Drainage Issue","Trap / Seal Issue","Contractor Non-compliance","Waterproofing Issue","Clamp / Support Issue","Pipe Damage","Other"]
STATUS_OPTIONS_ISSUE = ["Open","In Progress","Resolved","Closed"]

def ensure_issues_sheet():
    spreadsheet = client.open_by_url(SHEET_URL)
    try: return spreadsheet.worksheet(ISSUES_SHEET_NAME)
    except:
        ws = spreadsheet.add_worksheet(title=ISSUES_SHEET_NAME, rows=2000, cols=len(ISSUE_HEADERS))
        ws.append_row(ISSUE_HEADERS); return ws

@st.cache_data(ttl=60)
def load_issues():
    try:
        spreadsheet = client.open_by_url(SHEET_URL)
        try: ws = spreadsheet.worksheet(ISSUES_SHEET_NAME)
        except: return pd.DataFrame(columns=ISSUE_HEADERS)
        raw = ws.get_all_values()
        if not raw or len(raw)<2: return pd.DataFrame(columns=ISSUE_HEADERS)
        return clean_df(pd.DataFrame(raw[1:], columns=raw[0]))
    except Exception as e:
        st.error(f"Could not load issues: {e}"); return pd.DataFrame(columns=ISSUE_HEADERS)

def generate_issue_id(df):
    if df.empty or "Issue ID" not in df.columns: return "ISS-001"
    nums=[]
    for x in df["Issue ID"].dropna().tolist():
        try: nums.append(int(str(x).replace("ISS-","").strip()))
        except: pass
    return f"ISS-{str(max(nums)+1 if nums else 1).zfill(3)}"

def add_issue_to_sheet(row_data):
    try: ensure_issues_sheet().append_row(row_data); st.cache_data.clear(); return True
    except Exception as e: st.error(f"Failed to save issue: {e}"); return False

def update_issue_in_sheet(issue_id, new_status, resolution_notes):
    try:
        ws = client.open_by_url(SHEET_URL).worksheet(ISSUES_SHEET_NAME)
        raw = ws.get_all_values()
        if not raw: return False
        h = raw[0]
        try: ic=h.index("Issue ID"); sc=h.index("Status")+1; nc=h.index("Resolution Notes")+1; uc=h.index("Updated At")+1
        except: return False
        for rn, row in enumerate(raw[1:], start=2):
            if len(row)>ic and row[ic]==issue_id:
                ws.update_cell(rn,sc,new_status); ws.update_cell(rn,nc,resolution_notes)
                ws.update_cell(rn,uc,datetime.now().strftime("%d-%m-%Y %H:%M"))
                st.cache_data.clear(); return True
        return False
    except Exception as e: st.error(f"Failed to update issue: {e}"); return False

def build_issues_excel(issues_df):
    wb=Workbook(); ws=wb.active; ws.title="Site Issues"
    cols=[c for c in ["Issue ID","Site Name","Issue Type","Severity","Description","Status","Raised By","Raised Date","Assigned To","Target Date","Resolution Notes","Created At","Updated At"] if c in issues_df.columns]
    hf=PatternFill(start_color="111827",end_color="111827",fill_type="solid"); hfont=Font(color="FFFFFF",bold=True,size=11)
    tb=Border(left=Side(style="thin",color="D1D5DB"),right=Side(style="thin",color="D1D5DB"),top=Side(style="thin",color="D1D5DB"),bottom=Side(style="thin",color="D1D5DB"))
    for ci,cn in enumerate(cols,start=1):
        c=ws.cell(row=1,column=ci,value=cn); c.fill=hf; c.font=hfont; c.alignment=Alignment(horizontal="left",vertical="center"); c.border=tb
    sfm={"Open":PatternFill(start_color="FEE2E2",end_color="FEE2E2",fill_type="solid"),"In Progress":PatternFill(start_color="FEF3C7",end_color="FEF3C7",fill_type="solid"),"Resolved":PatternFill(start_color="DCFCE7",end_color="DCFCE7",fill_type="solid"),"Closed":PatternFill(start_color="E5E7EB",end_color="E5E7EB",fill_type="solid")}
    sff={"Open":Font(color="B91C1C",bold=True),"In Progress":Font(color="92400E",bold=True),"Resolved":Font(color="15803D",bold=True),"Closed":Font(color="374151",bold=True)}
    for ri,(_, row) in enumerate(issues_df[cols].astype(str).iterrows(),start=2):
        sv=row.get("Status","").strip(); rf=sfm.get(sv)
        for ci,cn in enumerate(cols,start=1):
            cel=ws.cell(row=ri,column=ci,value=row[cn]); cel.border=tb; cel.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
            if rf: cel.fill=rf
            if cn=="Status" and sv in sff: cel.font=sff[sv]
    wm={"Issue ID":10,"Site Name":22,"Issue Type":20,"Severity":10,"Description":40,"Status":14,"Raised By":16,"Raised Date":12,"Assigned To":16,"Target Date":12,"Resolution Notes":35,"Created At":16,"Updated At":16}
    for ci,cn in enumerate(cols,start=1): ws.column_dimensions[get_column_letter(ci)].width=wm.get(cn,16)
    ws.freeze_panes="A2"; ws.row_dimensions[1].height=22
    ws2=wb.create_sheet("Summary"); ws2.cell(row=1,column=1,value="Status").font=Font(bold=True); ws2.cell(row=1,column=2,value="Count").font=Font(bold=True)
    ws2.cell(row=1,column=1).fill=hf; ws2.cell(row=1,column=1).font=hfont; ws2.cell(row=1,column=2).fill=hf; ws2.cell(row=1,column=2).font=hfont
    if "Status" in issues_df.columns:
        for i,(sn,cv) in enumerate(issues_df["Status"].value_counts().items(),start=2):
            ws2.cell(row=i,column=1,value=sn); ws2.cell(row=i,column=2,value=int(cv))
            f=sfm.get(sn)
            if f: ws2.cell(row=i,column=1).fill=f; ws2.cell(row=i,column=2).fill=f
    ws2.column_dimensions["A"].width=18; ws2.column_dimensions["B"].width=12
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

def build_scanned_issues_excel(result_df):
    wb=Workbook(); ws=wb.active; ws.title="AI Detected Issues"
    col_map=[("site_name","Site Name"),("issue_type","Issue Type"),("severity","Severity"),("description","Description"),("raised_by","Raised By"),("raised_date","Date")]
    cols=[c for c,_ in col_map if c in result_df.columns]; hd=[l for c,l in col_map if c in result_df.columns]
    hf=PatternFill(start_color="111827",end_color="111827",fill_type="solid"); hfont=Font(color="FFFFFF",bold=True,size=11)
    tb=Border(left=Side(style="thin",color="D1D5DB"),right=Side(style="thin",color="D1D5DB"),top=Side(style="thin",color="D1D5DB"),bottom=Side(style="thin",color="D1D5DB"))
    for ci,label in enumerate(hd,start=1):
        c=ws.cell(row=1,column=ci,value=label); c.fill=hf; c.font=hfont; c.alignment=Alignment(horizontal="left",vertical="center"); c.border=tb
    sfm={"High":PatternFill(start_color="FEE2E2",end_color="FEE2E2",fill_type="solid"),"Medium":PatternFill(start_color="FEF3C7",end_color="FEF3C7",fill_type="solid"),"Low":PatternFill(start_color="DCFCE7",end_color="DCFCE7",fill_type="solid")}
    sff={"High":Font(color="B91C1C",bold=True),"Medium":Font(color="92400E",bold=True),"Low":Font(color="15803D",bold=True)}
    for ri,(_, row) in enumerate(result_df[cols].astype(str).iterrows(),start=2):
        sv=row.get("severity","").strip(); rf=sfm.get(sv)
        for ci,cn in enumerate(cols,start=1):
            cel=ws.cell(row=ri,column=ci,value=row[cn]); cel.border=tb; cel.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
            if rf: cel.fill=rf
            if cn=="severity" and sv in sff: cel.font=sff[sv]
    for ci,label in enumerate(hd,start=1): ws.column_dimensions[get_column_letter(ci)].width={"Site Name":22,"Issue Type":20,"Severity":10,"Description":45,"Raised By":16,"Date":12}.get(label,16)
    ws.freeze_panes="A2"; ws.row_dimensions[1].height=22
    buf=io.BytesIO(); wb.save(buf); buf.seek(0); return buf.getvalue()

def build_executive_pdf(display_df,tf,ts,tse,tp,sel_month,hc,hp,cg,summary_df):
    buf=io.BytesIO()
    doc=SimpleDocTemplate(buf,pagesize=landscape(A4),leftMargin=24,rightMargin=24,topMargin=24,bottomMargin=24)
    styles=getSampleStyleSheet()
    ts_=ParagraphStyle("T",parent=styles["Title"],fontSize=18,textColor=colors.HexColor("#111827"))
    ss_=ParagraphStyle("S",parent=styles["Normal"],fontSize=10,textColor=colors.HexColor("#4b5563"))
    hs_=ParagraphStyle("H",parent=styles["Heading3"],fontSize=12,textColor=colors.HexColor("#111827"),spaceBefore=10,spaceAfter=6)
    elems=[]
    elems.append(Paragraph("Huliot West Zone - Executive Dashboard Report",ts_))
    elems.append(Paragraph(f"Month: {sel_month}  |  Generated: {datetime.now().strftime('%d-%m-%Y %H:%M')}",ss_))
    elems.append(Spacer(1,14))
    kd=[["Total Floor Visits","Total Site Visits","Total Reports Sent","Total Pending Reports"],[str(tf),str(ts),str(tse),str(tp)]]
    kt=Table(kd,colWidths=[180,180,180,180])
    kt.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#111827")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTSIZE",(0,0),(-1,0),10),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,1),(-1,1),18),("FONTNAME",(0,1),(-1,1),"Helvetica-Bold"),("BACKGROUND",(0,1),(-1,1),colors.HexColor("#F9FAFB")),("TEXTCOLOR",(0,1),(-1,1),colors.HexColor("#111827")),("ALIGN",(0,0),(-1,-1),"CENTER"),("VALIGN",(0,0),(-1,-1),"MIDDLE"),("GRID",(0,0),(-1,-1),0.5,colors.HexColor("#D1D5DB")),("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8)]))
    elems.append(kt); elems.append(Spacer(1,16))
    if not summary_df.empty:
        elems.append(Paragraph("Reports Sent to Client - by Associate",hs_))
        cd=summary_df.sort_values("Report sent to the client",ascending=False)
        cats=[str(x)[:14] for x in cd["Associate ID"].tolist()]
        vals=[float(x) for x in cd["Report sent to the client"].tolist()]
        drw=Drawing(700,200); bc=VerticalBarChart()
        bc.x=50;bc.y=40;bc.height=140;bc.width=600;bc.data=[vals];bc.categoryAxis.categoryNames=cats
        bc.categoryAxis.labels.angle=30;bc.categoryAxis.labels.fontSize=7;bc.categoryAxis.labels.dx=-8;bc.categoryAxis.labels.dy=-10
        bc.valueAxis.valueMin=0;bc.bars[0].fillColor=colors.HexColor("#3b82f6")
        drw.add(bc); elems.append(drw); elems.append(Spacer(1,10))
    elems.append(Paragraph("Detailed Performance Breakdown",hs_))
    tc=[c for c in ["Associate ID","Floor Visit","Site Tower visit","Report Mark (YES)","Suggestion Visit (NO)","Report Pending","Report sent to the client"] if c in display_df.columns]
    td=[tc]+[list(r) for _,r in display_df[tc].astype(str).iterrows()]
    pt=Table(td,repeatRows=1,colWidths=[110]+[85]*(len(tc)-1))
    pc=[("BACKGROUND",(0,0),(-1,0),colors.HexColor("#111827")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),8),("ALIGN",(1,0),(-1,-1),"CENTER"),("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#D1D5DB")),("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5)]
    lri=len(td)-1; pc.append(("BACKGROUND",(0,lri),(-1,lri),colors.HexColor("#F3F4F6"))); pc.append(("FONTNAME",(0,lri),(-1,lri),"Helvetica-Bold"))
    pt.setStyle(TableStyle(pc)); elems.append(pt); elems.append(Spacer(1,16))
    elems.append(Paragraph("Highlights",hs_))
    ht=Table([["Highest Coverage",hc],["Highest Productivity",hp],["Critical Gaps",cg]],colWidths=[160,600])
    ht.setStyle(TableStyle([("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),9),("GRID",(0,0),(-1,-1),0.4,colors.HexColor("#D1D5DB")),("BACKGROUND",(0,0),(0,-1),colors.HexColor("#F9FAFB")),("TOPPADDING",(0,0),(-1,-1),6),("BOTTOMPADDING",(0,0),(-1,-1),6),("LEFTPADDING",(0,0),(-1,-1),8)]))
    elems.append(ht); doc.build(elems); buf.seek(0); return buf.getvalue()

def analyze_comments_for_issues(comment_records):
    api_key=st.secrets.get("gemini_api_key","")
    if not api_key: raise ValueError("gemini_api_key not found in Streamlit secrets.")
    valid=[r for r in comment_records if str(r.get("comment","")).strip() not in ["","-","nan","None"]]
    if not valid: return []
    prompt=f"""You are a senior plumbing site inspection analyst for Huliot Pipes & Fittings (West Zone India).
Analyze the following site visit comments. Extract ONLY real issues — defects, non-compliances, problems, pending work, or observations that need action.
Skip: general progress updates, positive comments, routine check-ins with no problems.
Visit Comments:\n{json.dumps(valid,indent=2,ensure_ascii=False)}
Return a JSON array. Each item:
{{"site_name":"exact site name from the record","issue_type":"one of: Installation Defect | Material Issue | Design Non-compliance | Slope / Drainage Issue | Trap / Seal Issue | Contractor Non-compliance | Waterproofing Issue | Clamp / Support Issue | Pipe Damage | Other","severity":"High | Medium | Low","description":"clear 1-2 sentence description","raised_by":"associate name from the record","raised_date":"date from the record"}}
Rules: One issue per finding. High=safety risk/major defect. Medium=non-compliance/rework. Low=minor/cosmetic.
Return ONLY valid JSON array, no other text, no markdown fences."""
    url=f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
    payload={"contents":[{"parts":[{"text":prompt}]}],"generationConfig":{"temperature":0.2,"maxOutputTokens":8192,"responseMimeType":"application/json"}}
    r=requests.post(url,headers={"content-type":"application/json"},json=payload,timeout=60)
    d=r.json()
    if r.status_code!=200: raise Exception(d.get("error",{}).get("message","Gemini API error"))
    try:
        cand=d["candidates"][0]
        raw=cand["content"]["parts"][0]["text"].strip()
    except(KeyError,IndexError): raise Exception("Unexpected Gemini response format.")
    if cand.get("finishReason","")=="MAX_TOKENS": raise Exception("Gemini response cut off — filter to fewer comments and retry.")
    if raw.startswith("```"):
        raw=raw.split("```")[1]
        if raw.startswith("json"): raw=raw[4:]
    return json.loads(raw.strip())

# ==========================================
# GEOCODING — CITY + AREA (NEIGHBORHOOD) LEVEL
# Used by Site Map tab. No API key needed.
# ==========================================

# City-level centroids
CITY_COORDS = {
    # Maharashtra
    "mumbai":(19.0760,72.8777),"pune":(18.5204,73.8567),"nagpur":(21.1458,79.0882),
    "nashik":(19.9975,73.7898),"thane":(19.2183,72.9781),"aurangabad":(19.8762,75.3433),
    "chhatrapati sambhajinagar":(19.8762,75.3433),"solapur":(17.6599,75.9064),
    "kolhapur":(16.7050,74.2433),"amravati":(20.9320,77.7523),"navi mumbai":(19.0330,73.0297),
    "vasai":(19.4912,72.8054),"virar":(19.4559,72.8112),"kalyan":(19.2403,73.1305),
    "dombivli":(19.2183,73.0864),"panvel":(18.9894,73.1175),"ahmednagar":(19.0948,74.7480),
    "jalgaon":(21.0077,75.5626),"akola":(20.7002,77.0082),"latur":(18.4088,76.5604),
    "dhule":(20.9042,74.7749),"ratnagiri":(16.9902,73.3120),"satara":(17.6805,74.0183),
    "sangli":(16.8524,74.5815),"wardha":(20.7453,78.6022),"chandrapur":(19.9615,79.2961),
    # Gujarat
    "ahmedabad":(23.0225,72.5714),"surat":(21.1702,72.8311),"vadodara":(22.3072,73.1812),
    "baroda":(22.3072,73.1812),"rajkot":(22.3039,70.8022),"bhavnagar":(21.7645,72.1519),
    "jamnagar":(22.4707,70.0577),"gandhinagar":(23.2156,72.6369),"anand":(22.5645,72.9289),
    "navsari":(20.9467,72.9520),"morbi":(22.8173,70.8378),"bharuch":(21.7051,72.9959),
    "mehsana":(23.5880,72.3693),"vapi":(20.3893,72.9106),"valsad":(20.5992,72.9342),
    # Madhya Pradesh
    "indore":(22.7196,75.8577),"bhopal":(23.2599,77.4126),"jabalpur":(23.1815,79.9864),
    "gwalior":(26.2183,78.1828),"ujjain":(23.1765,75.7885),"dewas":(22.9676,76.0534),
    "ratlam":(23.3315,75.0367),"satna":(24.5854,80.8322),"rewa":(24.5362,81.2961),
    "burhanpur":(21.3009,76.2291),"khandwa":(21.8245,76.3529),"chhindwara":(22.0574,78.9382),
    "sagar":(23.8388,78.7378),"katni":(23.8339,80.3933),
    # Chhattisgarh
    "raipur":(21.2514,81.6296),"bhilai":(21.2090,81.4285),"bilaspur":(22.0797,82.1391),
    "korba":(22.3595,82.7501),"durg":(21.1904,81.2849),"jagdalpur":(19.0822,82.0322),
    "rajnandgaon":(21.0974,81.0388),"raigarh":(21.8974,83.3950),"ambikapur":(23.1193,83.1957),
    # Other India
    "delhi":(28.7041,77.1025),"new delhi":(28.6139,77.2090),
    "bengaluru":(12.9716,77.5946),"bangalore":(12.9716,77.5946),
    "hyderabad":(17.3850,78.4867),"chennai":(13.0827,80.2707),
    "kolkata":(22.5726,88.3639),"jaipur":(26.9124,75.7873),
    "lucknow":(26.8467,80.9462),"patna":(25.5941,85.1376),
    "ranchi":(23.3441,85.3096),"bhubaneswar":(20.2961,85.8245),
    "chandigarh":(30.7333,76.7794),"dehradun":(30.3165,78.0322),
    "panaji":(15.4909,73.8278),"guwahati":(26.1445,91.7362),
}

# Neighborhood / area level — gives precise placement inside cities
AREA_COORDS = {
    # ── PUNE ──────────────────────────────────────────
    "baner":(18.5590,73.7868),"wakad":(18.5994,73.7614),"kothrud":(18.5074,73.8077),
    "aundh":(18.5626,73.8081),"viman nagar":(18.5642,73.9142),"koregaon park":(18.5362,73.8938),
    "hadapsar":(18.5018,73.9395),"kharadi":(18.5531,73.9446),"magarpatta":(18.5133,73.9290),
    "pimple saudagar":(18.6196,73.8000),"hinjawadi":(18.5908,73.7390),"pashan":(18.5297,73.8013),
    "bawdhan":(18.5264,73.7642),"bavdhan":(18.5264,73.7642),"undri":(18.4616,73.9034),
    "kondhwa":(18.4628,73.8868),"wanowrie":(18.4944,73.9032),"mundhwa":(18.5296,73.9323),
    "bibwewadi":(18.4713,73.8468),"katraj":(18.4509,73.8626),"ambegaon":(18.4646,73.8525),
    "sus":(18.5426,73.7477),"mahalunge":(18.5805,73.7244),"pimple nilakh":(18.6116,73.8067),
    "ravet":(18.6482,73.7527),"punawale":(18.6295,73.7536),"chinchwad":(18.6447,73.8024),
    "pimpri":(18.6253,73.7979),"nigdi":(18.6651,73.7736),"talegaon":(18.7325,73.6767),
    "chakan":(18.7630,73.8612),"shivajinagar":(18.5308,73.8474),"deccan":(18.5161,73.8416),
    "camp pune":(18.5155,73.8803),"yerwada":(18.5540,73.8917),"dhanori":(18.5769,73.9195),
    "lohegaon":(18.5962,73.9125),"manjri":(18.5065,73.9575),"wagholi":(18.5724,73.9826),
    "nanded city":(18.4547,73.8067),"pirangut":(18.5083,73.7275),"urse":(18.6511,73.7196),
    "moshi":(18.6794,73.8533),"alandi":(18.6736,73.8939),"tathawade":(18.6091,73.7686),
    "akurdi":(18.6476,73.7617),"sangvi":(18.6085,73.8033),"kiwale":(18.6344,73.7397),
    "pimple gurav":(18.6103,73.7896),"mamurdi":(18.6229,73.7507),"bhukum":(18.5148,73.7419),
    "talawade":(18.6340,73.7600),"dehu road":(18.6837,73.7591),"narhe":(18.4680,73.8230),
    "dhayari":(18.4574,73.8208),"vadgaon sheri":(18.5521,73.9258),"fursungi":(18.4757,73.9327),
    "uruli kanchan":(18.4386,74.0416),"sinhagad road":(18.4760,73.8175),
    "kesnand":(18.5400,73.9800),"nandoshi":(18.6900,73.8200),"kudalwadi":(18.6560,73.7890),
    "shirgaon":(18.6100,73.7700),"dehu":(18.7119,73.7558),"khed shivapur":(18.3600,73.8400),
    "ambil odha":(18.4900,73.8100),"vadu":(18.7000,74.0000),"bakori":(18.5000,74.1000),
    "saswad":(18.3447,74.0214),"lonavala":(18.7481,73.4072),"khandala":(18.7600,73.3700),
    "lavasa":(18.3980,73.5100),"mulshi":(18.5100,73.5300),"baramati":(18.1539,74.5795),
    "jejuri":(18.2749,74.1548),"bhor":(18.1572,73.8469),"shirur":(18.8263,74.3636),
    "rajgurunagar":(18.8417,73.9952),"talegaon dabhade":(18.7325,73.6767),
    # ── NAVI MUMBAI ───────────────────────────────────
    "kharghar":(19.0474,73.0677),"belapur":(19.0225,73.0390),"vashi":(19.0771,73.0097),
    "airoli":(19.1561,72.9979),"ghansoli":(19.1294,73.0008),"kopar khairane":(19.1061,73.0074),
    "nerul":(19.0328,73.0159),"seawoods":(19.0152,73.0196),"ulwe":(18.9636,73.0552),
    "dronagiri":(18.9456,72.9954),"taloja":(19.0040,73.1193),"kamothe":(19.0341,73.0699),
    "kalamboli":(19.0213,73.0891),"roadpali":(19.0520,73.0895),"cbd belapur":(19.0213,73.0296),
    "nhava sheva":(18.9550,72.9431),"panvel":(18.9894,73.1175),
    # ── MUMBAI ────────────────────────────────────────
    "andheri east":(19.1136,72.8697),"andheri west":(19.1266,72.8370),"andheri":(19.1136,72.8697),
    "bandra west":(19.0596,72.8295),"bandra east":(19.0654,72.8459),"bandra":(19.0596,72.8295),
    "bandra kurla":(19.0660,72.8654),"bkc":(19.0660,72.8654),"khar":(19.0726,72.8361),
    "borivali east":(19.2307,72.8567),"borivali west":(19.2230,72.8468),"borivali":(19.2307,72.8567),
    "kandivali east":(19.2042,72.8491),"kandivali west":(19.2042,72.8352),"kandivali":(19.2042,72.8491),
    "malad east":(19.1874,72.8483),"malad west":(19.1874,72.8353),"malad":(19.1874,72.8483),
    "goregaon east":(19.1663,72.8526),"goregaon west":(19.1605,72.8448),"goregaon":(19.1663,72.8526),
    "jogeshwari east":(19.1317,72.8488),"jogeshwari west":(19.1278,72.8383),"jogeshwari":(19.1317,72.8488),
    "santacruz east":(19.0822,72.8556),"santacruz west":(19.0822,72.8383),"santacruz":(19.0822,72.8383),
    "vile parle east":(19.1064,72.8613),"vile parle west":(19.0993,72.8479),"vile parle":(19.0993,72.8479),
    "dadar":(19.0178,72.8478),"worli":(19.0176,72.8162),"kurla":(19.0665,72.8790),
    "ghatkopar":(19.0868,72.9081),"vikhroli":(19.1084,72.9258),"mulund":(19.1759,72.9573),
    "bhandup":(19.1535,72.9360),"powai":(19.1218,72.9052),"chembur":(19.0622,72.8990),
    "dharavi":(19.0422,72.8527),"sion":(19.0396,72.8636),"matunga":(19.0277,72.8497),
    "parel":(19.0028,72.8384),"lower parel":(18.9956,72.8236),"byculla":(18.9780,72.8324),
    "grant road":(18.9649,72.8181),"marine lines":(18.9455,72.8245),"colaba":(18.9067,72.8147),
    "fort":(18.9322,72.8354),"churchgate":(18.9352,72.8265),"nariman point":(18.9247,72.8235),
    "oshiwara":(19.1420,72.8346),"versova":(19.1373,72.8115),"lokhandwala":(19.1381,72.8312),
    "four bungalows":(19.1196,72.8466),"seven bungalows":(19.1310,72.8345),
    "mindspace":(19.1589,72.8530),"nesco":(19.1447,72.8530),"peninsula corporate":(19.1206,72.8679),
    "ghodbunder road":(19.2564,72.9717),"majiwada":(19.2054,72.9755),"manpada":(19.1871,73.0072),
    "hiranandani estate":(19.2614,72.9847),"brahmand":(19.2780,73.0063),
    "dombivli east":(19.2183,73.0864),"dombivli west":(19.2172,73.0704),"dombivli":(19.2183,73.0864),
    "kalyan east":(19.2403,73.1305),"kalyan west":(19.2513,73.1262),"kalyan":(19.2403,73.1305),
    "titwala":(19.3137,73.2006),"badlapur":(19.1619,73.2524),"ambernath":(19.1967,73.1916),
    "ulhasnagar":(19.2187,73.1638),"mumbra":(19.1900,73.0150),"diva":(19.1960,73.0500),
    "thane west":(19.1974,72.9782),"thane east":(19.2162,73.0170),
    "kapurbawdi":(19.1980,72.9750),"vartak nagar":(19.2050,72.9650),"balkum":(19.1850,72.9650),
    "naupada":(19.1850,72.9750),"uthalsar":(19.1900,72.9900),"kolshet":(19.2150,72.9680),
    "waghbil":(19.2690,72.9950),"pokhran road":(19.2300,72.9900),
    "vasai west":(19.4912,72.8054),"vasai east":(19.4740,72.8396),"vasai":(19.4912,72.8054),
    "virar west":(19.4559,72.8112),"virar east":(19.4630,72.8300),"virar":(19.4559,72.8112),
    "nalasopara west":(19.4146,72.7990),"nalasopara east":(19.4040,72.8200),"nalasopara":(19.4146,72.7990),
    "mira road":(19.2972,72.8709),"bhayandar":(19.2990,72.8504),
    # ── AHMEDABAD ─────────────────────────────────────
    "sg highway":(23.0397,72.5097),"satellite ahmedabad":(23.0303,72.5254),"satellite":(23.0303,72.5254),
    "prahladnagar":(23.0068,72.5087),"prahlad nagar":(23.0068,72.5087),
    "bopal":(23.0347,72.4631),"south bopal":(23.0100,72.4730),"ambli":(23.0500,72.4800),
    "shela":(22.9960,72.4600),"ghuma":(23.0100,72.4850),"sarkhej":(23.0050,72.5000),
    "gota":(23.1196,72.5501),"chandkheda":(23.1043,72.5825),"motera":(23.1047,72.6042),
    "sabarmati":(23.0853,72.5824),"ranip":(23.0699,72.5725),"nava vadaj":(23.0600,72.5900),
    "gurukul":(23.0650,72.5400),"thaltej":(23.0497,72.4988),"bodakdev":(23.0474,72.5109),
    "vastrapur":(23.0359,72.5257),"navrangpura":(23.0336,72.5595),"ellisbridge":(23.0148,72.5697),
    "ambawadi":(23.0222,72.5530),"iscon ahmedabad":(23.0352,72.5047),
    "drive in":(23.0523,72.5333),"memnagar":(23.0586,72.5394),"ghatlodia":(23.0717,72.5379),
    "naranpura":(23.0544,72.5580),"maninagar":(22.9964,72.6097),"vastral":(23.0340,72.6645),
    "naroda":(23.0712,72.6536),"odhav":(22.9999,72.6469),"vatva":(22.9560,72.6399),
    "rakhial":(23.0377,72.6313),"vejalpur":(22.9904,72.5536),"jivraj park":(23.0001,72.5520),
    "anandnagar":(23.0280,72.5250),"manekbaug":(23.0190,72.5459),"vasna ahmedabad":(23.0007,72.5456),
    "paldi":(23.0094,72.5629),"shahibaug":(23.0578,72.5960),"kankaria":(23.0041,72.5980),
    "cg road":(23.0335,72.5656),"ashram road":(23.0337,72.5711),"science city":(23.0850,72.5250),
    "bhat":(23.1350,72.5900),"shilaj":(23.0600,72.4700),"hebatpur":(23.1000,72.5100),
    # ── SURAT ─────────────────────────────────────────
    "adajan":(21.2158,72.7996),"vesu":(21.1481,72.7772),"pal surat":(21.1800,72.8000),
    "althan":(21.1640,72.7830),"dumas":(21.1010,72.7390),"ghod dod":(21.1920,72.8210),
    "athwa":(21.1960,72.8120),"citylight":(21.1810,72.7900),"katargam":(21.2400,72.8400),
    "varachha":(21.2070,72.8700),"kapodra":(21.2050,72.8550),"rander":(21.2590,72.7930),
    "sachin":(21.0890,72.8830),"dindoli":(21.1600,72.8600),"nana varachha":(21.2070,72.8700),
    "bhesan":(21.1810,72.7680),"piplod":(21.2100,72.7900),"bhatar":(21.2300,72.8400),
    # ── VADODARA ──────────────────────────────────────
    "alkapuri":(22.3217,73.1741),"sayajigunj":(22.3072,73.1812),"akota":(22.2900,73.1680),
    "manjalpur":(22.2630,73.1860),"gotri":(22.3353,73.1579),"waghodia road":(22.3600,73.2300),
    "tandalja":(22.2990,73.1560),"fatehgunj":(22.3220,73.1950),"harni":(22.3200,73.2200),
    "sama":(22.3430,73.2150),"tarsali":(22.2450,73.2100),"makarpura":(22.2600,73.1900),
    "danteshwar":(22.2700,73.2300),"karelibaug":(22.3200,73.2000),"nizampura":(22.3400,73.2100),
    # ── NAGPUR ────────────────────────────────────────
    "sitabuldi":(21.1498,79.0882),"dharampeth":(21.1363,79.0667),"ramdaspeth":(21.1404,79.0754),
    "sadar nagpur":(21.1518,79.0756),"civil lines nagpur":(21.1634,79.0891),
    "gandhibagh":(21.1457,79.0960),"lakadganj":(21.1308,79.0970),"mankapur":(21.1136,79.0758),
    "nandanvan":(21.1045,79.1125),"wathoda":(21.0858,79.1194),"ajni":(21.1145,79.1001),
    "itwari":(21.1488,79.1088),"hingna":(21.1026,78.9748),"butibori":(21.0000,79.0000),
    "khamla":(21.1200,79.0550),"sakkardara":(21.1150,79.1050),"trimurti nagar":(21.1280,79.0870),
    "manish nagar":(21.1060,79.0950),"laxmi nagar nagpur":(21.1380,79.0840),
    "jaripatka":(21.1300,79.1100),"pachpaoli":(21.1200,79.0900),"kalamna":(21.1350,79.1200),
    "beltarodi":(21.0750,79.0150),"khapri":(21.0880,79.1200),
    "wardha road nagpur":(21.1000,79.0500),"amravati road nagpur":(21.1900,78.9300),
    "bhandara road nagpur":(21.1600,79.1400),
    # ── NASHIK ────────────────────────────────────────
    "gangapur road":(20.0224,73.7842),"college road nashik":(20.0036,73.7844),
    "panchavati nashik":(20.0063,73.7929),"satpur":(19.9870,73.7430),
    "ambad nashik":(19.9755,73.7577),"pathardi phata":(20.0100,73.8200),
    "cidco nashik":(19.9700,73.8300),"dwarka nashik":(19.9500,73.8400),
    "mhasrul":(20.0296,73.7644),"trimbak road":(19.9600,73.7200),
    "indira nagar nashik":(20.0020,73.7700),"igatpuri":(19.6940,73.5610),
    # ── THANE ─────────────────────────────────────────
    "hiranandani":(19.2614,72.9847),"ghodbunder":(19.2564,72.9717),
    "teen haath naka":(19.2050,72.9700),"pokhran":(19.2300,72.9900),"owale":(19.2750,72.9720),
    # ── INDORE ────────────────────────────────────────
    "vijay nagar indore":(22.7319,75.8879),"palasia":(22.7198,75.8728),"rau":(22.6431,75.8415),
    "lasudia":(22.7550,75.9003),"super corridor":(22.7853,75.9053),"nipania":(22.7464,75.9183),
    "bhawarkuan":(22.7080,75.8476),"geeta bhawan":(22.7225,75.8574),"annapurna":(22.6938,75.8598),
    "bicholi mardana":(22.7600,75.9500),"scheme 54":(22.7300,75.9000),"scheme 78":(22.7600,75.9000),
    "aerodrome road":(22.7270,75.8017),"kanadiya":(22.7800,75.8800),"khajrana":(22.6770,75.8990),
    "sanwer road":(22.7300,75.7900),"dewas road indore":(22.7500,75.9400),
    "mhow":(22.5583,75.7709),"pithampur":(22.6169,75.6944),"musakhedi":(22.7100,75.8700),
    "pipliyahana":(22.6900,75.8900),"rao rampura":(22.6850,75.8700),
    # ── BHOPAL ────────────────────────────────────────
    "mp nagar":(23.2306,77.4169),"arera colony":(23.2153,77.4313),
    "hoshangabad road":(23.2100,77.4600),"kolar road":(23.1900,77.4400),
    "ayodhya nagar":(23.2497,77.3985),"chuna bhatti":(23.2200,77.4500),
    "katara hills":(23.2700,77.3900),"misrod":(23.2800,77.4700),"bairagarh":(23.2750,77.3550),
    "mandideep":(23.1000,77.5100),"berasia road":(23.3100,77.4300),
    # ── RAIPUR ────────────────────────────────────────
    "shankar nagar":(21.2615,81.6340),"telibandha":(21.2447,81.6317),
    "civil lines raipur":(21.2699,81.6282),"pachpedi naka":(21.2231,81.6446),
    "pandri":(21.2600,81.6600),"avanti vihar":(21.2300,81.6400),
    "tatibandh":(21.2650,81.6950),"mowa raipur":(21.3100,81.6300),
    "vidhan sabha road":(21.2500,81.6200),"bhatagaon":(21.2700,81.7400),
    # ── BHILAI ────────────────────────────────────────
    "supela":(21.2100,81.4100),"smriti nagar":(21.2000,81.4300),"risali":(21.2200,81.3800),
    "nehru nagar bhilai":(21.1900,81.3900),"hudco":(21.2300,81.4100),
    "vaishali nagar":(21.2050,81.4400),"ruabandha":(21.2400,81.4600),"charoda":(21.2600,81.3800),
    # ── BILASPUR ──────────────────────────────────────
    "vyapaar vihar":(22.0800,82.1400),"torwa":(22.0650,82.1200),"link road bilaspur":(22.0750,82.1350),
    "mangla bilaspur":(22.0900,82.1600),"sarkanda":(22.1000,82.1500),
    "seepat road":(22.1200,82.2000),"koni":(22.0400,82.1200),
}

STATE_COORDS = {
    "maharashtra":(19.7515,75.7139),"gujarat":(22.2587,71.1924),
    "madhya pradesh":(22.9734,78.6569),"chhattisgarh":(21.2787,81.8661),
    "rajasthan":(27.0238,74.2179),"karnataka":(15.3173,75.7139),
    "tamil nadu":(11.1271,78.6569),"telangana":(18.1124,79.0193),
    "andhra pradesh":(15.9129,79.7400),"uttar pradesh":(26.8467,80.9462),
    "delhi":(28.7041,77.1025),"punjab":(31.1471,75.3412),
    "haryana":(29.0588,76.0856),"west bengal":(22.9868,87.8550),
    "bihar":(25.0961,85.3131),"kerala":(10.8505,76.2711),
    "odisha":(20.9517,85.0985),"jharkhand":(23.6102,85.2799),
    "goa":(15.2993,74.1240),"assam":(26.2006,92.9376),
    "uttarakhand":(30.0668,79.0193),"himachal pradesh":(31.1048,77.1734),
}

import re as _re

# Noise words common in project names — strip before area matching
_NOISE = _re.compile(
    r'\b(project|residency|residencies|heights|park|city|town|township|nagar|'
    r'society|complex|tower|towers|building|plaza|mall|sector|phase|block|'
    r'apartment|apartments|enclave|estate|garden|gardens|villa|villas|'
    r'homes|home|residences|residence|infra|developers|builders|construction|'
    r'pvt|ltd|private|limited|properties|property|real|realty|group|'
    r'skyline|horizon|landmark|signature|grand|royal|imperial|premium|'
    r'one|two|three|four|five|1|2|3|4|5|a|b|c|d)\b',
    _re.IGNORECASE
)

def geocode_site(project_name, city_or_district, state):
    """
    3-level precision with word-boundary matching (no substring false-positives).
    Priority: area neighborhood > city/district > state centroid.
    Returns (lat, lon, level)
    """
    pn_raw = str(project_name).lower()
    c_raw  = str(city_or_district).lower().strip()
    s_raw  = str(state).lower().strip()

    # Strip noise so "Baner Heights Pvt Ltd" becomes "baner"
    pn_clean = _re.sub(r'\s+', ' ', _NOISE.sub(' ', pn_raw)).strip()

    # Unified search space
    search = pn_clean + ' ' + c_raw + ' ' + pn_raw

    # ── Level 1: AREA (word-boundary, longest match first) ─────────────
    for key, (lat, lon) in sorted(AREA_COORDS.items(), key=lambda x: -len(x[0])):
        pat = r'(?<![a-z])' + _re.escape(key) + r'(?![a-z])'
        if _re.search(pat, search):
            return lat, lon, 'area'

    # ── Level 2: CITY (exact first, then word-boundary) ─────────────────
    if c_raw in CITY_COORDS:
        return CITY_COORDS[c_raw][0], CITY_COORDS[c_raw][1], 'city'
    for key, (lat, lon) in sorted(CITY_COORDS.items(), key=lambda x: -len(x[0])):
        pat = r'(?<![a-z])' + _re.escape(key) + r'(?![a-z])'
        if c_raw and _re.search(pat, c_raw):
            return lat, lon, 'city'

    # ── Level 3: STATE ───────────────────────────────────────────────────
    if s_raw in STATE_COORDS:
        return STATE_COORDS[s_raw][0], STATE_COORDS[s_raw][1], 'state'
    for key, (lat, lon) in STATE_COORDS.items():
        pat = r'(?<![a-z])' + _re.escape(key) + r'(?![a-z])'
        if s_raw and _re.search(pat, s_raw):
            return lat, lon, 'state'

    return None, None, None

# ==========================================
# LOAD DATA
# ==========================================
@st.cache_data(ttl=120)
def load_data():
    try: spreadsheet = client.open_by_url(SHEET_URL)
    except Exception as e: st.error(f"Could not open Google Sheet: {e}"); return pd.DataFrame(), pd.DataFrame()
    visit_dfs=[]; master_df=pd.DataFrame()
    for ws in spreadsheet.worksheets():
        title=ws.title.lower()
        raw=ws.get_all_values()
        if not raw or len(raw)<2: continue
        df=clean_df(pd.DataFrame(raw[1:],columns=make_unique_headers(raw[0])))
        if "master" in title: master_df=df; master_df["Source Sheet"]=ws.title; continue
        if any(skip in title for skip in ["setting","config","associate",ISSUES_SHEET_NAME.lower()]): continue
        if not df.empty and ("Site Name" in df.columns or "Visit ID" in df.columns):
            df["Source Sheet"]=ws.title; visit_dfs.append(df)
    visits_df=clean_df(pd.concat(visit_dfs,ignore_index=True)) if visit_dfs else pd.DataFrame()
    master_df=clean_df(master_df)
    return visits_df, master_df

visits_df, master_df = load_data()
issues_df = load_issues()

with st.sidebar:
    if st.button("🔄 Refresh Data"):
        st.cache_data.clear(); st.rerun()
    st.caption(f"🕐 Auto-refreshes every 2 min\nLast load: {datetime.now().strftime('%H:%M:%S')}")
    st.markdown("---")

if not visits_df.empty:
    visits_df["Status"]=visits_df.apply(get_visit_status,axis=1)
    dc=safe_col(visits_df,["Date of Visit","Visit Date","Date"])
    if dc:
        visits_df["Date Parsed"]=pd.to_datetime(visits_df[dc],errors="coerce")
        visits_df["Month"]=visits_df["Date Parsed"].dt.strftime("%b %Y").fillna("Unknown")
    else:
        visits_df["Date Parsed"]=pd.NaT; visits_df["Month"]="Unknown"
    fc=safe_col(visits_df,["FloorsVisited","Floors Visited","Floor Visited","Floor"])
    visits_df["Num_Floors"]=visits_df[fc].apply(parse_floor) if fc else 0
    visits_df["Clean_Report_Mark"]=visits_df["Is Report Visit?"].astype(str).str.strip().str.upper() if "Is Report Visit?" in visits_df.columns else ""

# ==========================================
# UI
# ==========================================
st.title("📊 Site Visit Deep Analytics")
st.markdown("Live data synchronized directly from your Google Sheets.")

tab_visits,tab_master,tab_exec,tab_site_card,tab_issues,tab_map = st.tabs([
    "📊 Visit Analytics","📈 Master Projects","👔 Executive Dashboard",
    "🏢 Site Report Card","🚨 Site Issues","🗺️ Site Map"
])

# ==========================================
# TAB 1: VISIT ANALYTICS
# ==========================================
with tab_visits:
    if visits_df.empty:
        st.warning("No Visit Log data found.")
    else:
        st.subheader("Data Filters")
        c1,c2,c3,c4,c5=st.columns(5)
        ac=safe_col(visits_df,["Associate ID","Associate","Technical Person"])
        sc=find_visit_site_col(visits_df)
        f_src=c1.selectbox("Source Sheet",["All"]+clean_options(visits_df["Source Sheet"]),key="t1_source")
        f_mon=c2.selectbox("Month",["All"]+clean_options(visits_df["Month"]),key="t1_month")
        f_sta=c3.selectbox("Status",["All"]+clean_options(visits_df["Status"]),key="t1_status")
        f_asc=c4.selectbox("Associate",["All"]+(clean_options(visits_df[ac]) if ac else []),key="t1_assoc")
        f_sit=c5.selectbox("Site Name",["All"]+(clean_options(visits_df[sc]) if sc else []),key="t1_site")
        fv=visits_df.copy()
        if f_src!="All": fv=fv[fv["Source Sheet"].astype(str)==f_src]
        if f_mon!="All": fv=fv[fv["Month"].astype(str)==f_mon]
        if f_sta!="All": fv=fv[fv["Status"].astype(str)==f_sta]
        if ac and f_asc!="All": fv=fv[fv[ac].astype(str)==f_asc]
        if sc and f_sit!="All": fv=fv[fv[sc].astype(str)==f_sit]
        k1,k2,k3,k4,k5=st.columns(5)
        k1.metric("Total Visits",int(fv["Num_Floors"].sum()))
        k2.metric("Pending Reports",len(fv[fv["Status"]=="Pending"]))
        k3.metric("Technical (NA)",int(fv[fv["Status"]=="Technical (NA)"]["Num_Floors"].sum()))
        k4.metric("Submitted",len(fv[fv["Status"]=="Submitted"]))
        k5.metric("Submitted Floors",int(fv[fv["Clean_Report_Mark"].isin(["YES","Y","TRUE"])]["Num_Floors"].sum()))
        st.markdown("---")
        ch1,ch2=st.columns(2)
        with ch1:
            with st.container(border=True):
                st.markdown("##### Visits Per Month")
                mc=fv["Month"].value_counts().reset_index(); mc.columns=["Month","Visits"]
                st.plotly_chart(px.bar(mc,x="Month",y="Visits",color_discrete_sequence=["#6366f1"]),use_container_width=True,key="c_t1m")
        with ch2:
            with st.container(border=True):
                st.markdown("##### Top Sites / Zones")
                if sc:
                    sic=fv[sc].value_counts().nlargest(6).reset_index(); sic.columns=["Site Name","Visits"]
                    st.plotly_chart(px.pie(sic,names="Site Name",values="Visits",hole=0.4,color_discrete_sequence=["#6366f1","#14b8a6","#f59e0b","#f43f5e","#8b5cf6","#0ea5e9"]),use_container_width=True,key="c_t1p")
        st.subheader("Visit Records")
        dcols=[c for c in ["Source Sheet","Visit ID",sc,"Tower Name","FloorsVisited","Floors Visited",ac,"Date of Visit","Status","Report Submitted Date","Comment"] if c and c in fv.columns]
        dcols=list(dict.fromkeys(dcols))
        st.dataframe(fv[dcols].astype(str),use_container_width=True,hide_index=True)
        st.markdown("---")
        st.markdown("#### 🤖 AI Comment Analyzer")
        st.caption("Scans filtered visit comments and detects potential site issues via Gemini AI.")
        cc_=safe_col(fv,["Comment","Remarks","Observation"])
        dc_=safe_col(fv,["Date of Visit","Visit Date","Date"])
        ac_=safe_col(fv,["Associate ID","Associate","Technical Person"])
        sc_=find_visit_site_col(fv)
        if not cc_:
            st.warning("No Comment column found.")
        else:
            hc=fv[fv[cc_].astype(str).str.strip().isin(["","-","nan","None"])==False]
            nc=len(hc)
            col_s,col_c=st.columns([2,1])
            with col_s:
                scan_btn=st.button(f"🔍 Scan {nc} Comments for Issues",key="btn_scan",disabled=(nc==0))
            with col_c:
                if st.button("🗑️ Clear Results",key="btn_clear_ai"):
                    st.session_state["analyzed_issues"]=[]; st.rerun()
            if scan_btn:
                recs=[{"site_name":str(r.get(sc_,"Unknown")) if sc_ else "Unknown","date":str(r.get(dc_,"")) if dc_ else "","associate":str(r.get(ac_,"")) if ac_ else "","comment":str(r.get(cc_,"")).strip()} for _,r in hc.iterrows()]
                with st.spinner(f"Gemini analyzing {nc} comments..."):
                    try:
                        det=analyze_comments_for_issues(recs); st.session_state["analyzed_issues"]=det
                        if not det: st.info("✅ No actionable issues detected.")
                    except json.JSONDecodeError: st.error("⚠️ Gemini returned malformed JSON — filter to fewer comments and retry."); st.session_state["analyzed_issues"]=[]
                    except ValueError as ve: st.error(f"⚠️ {ve}"); st.session_state["analyzed_issues"]=[]
                    except Exception as ex: st.error(f"Analysis failed: {ex}"); st.session_state["analyzed_issues"]=[]
            analyzed=st.session_state.get("analyzed_issues",[])
            if analyzed:
                st.success(f"🔎 Detected **{len(analyzed)} issues**. Review and select which to add.")
                rdf=pd.DataFrame(analyzed)
                for cn in ["site_name","issue_type","severity","description","raised_by","raised_date"]:
                    if cn not in rdf.columns: rdf[cn]=""
                rdf.insert(0,"Add?",True)
                edf=st.data_editor(rdf,column_config={"Add?":st.column_config.CheckboxColumn("Add?",default=True,width="small"),"site_name":st.column_config.TextColumn("Site Name",width="medium"),"issue_type":st.column_config.TextColumn("Issue Type",width="medium"),"severity":st.column_config.SelectboxColumn("Severity",options=["High","Medium","Low"],width="small"),"description":st.column_config.TextColumn("Description",width="large"),"raised_by":st.column_config.TextColumn("Raised By",width="small"),"raised_date":st.column_config.TextColumn("Date",width="small")},use_container_width=True,hide_index=True,key="ai_editor")
                sr_=edf[edf["Add?"]==True]; ns_=len(sr_)
                dx1,dx2=st.columns([1,3])
                with dx1:
                    st.download_button("📥 Download Excel",data=build_scanned_issues_excel(edf.drop(columns=["Add?"])),file_name=f"AI_Issues_{datetime.now().strftime('%d%m%Y')}.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",key="dl_scan_xl")
                with dx2:
                    st.caption("🔴 High  🟡 Medium  🟢 Low — color-coded by severity.")
                if ns_>0:
                    if st.button(f"➕ Add {ns_} Selected → Issue Tracker",key="btn_add_ai",type="primary"):
                        fi_=load_issues(); en_=[]
                        if not fi_.empty and "Issue ID" in fi_.columns:
                            for x in fi_["Issue ID"].dropna():
                                try: en_.append(int(str(x).replace("ISS-","").strip()))
                                except: pass
                        sn_=max(en_)+1 if en_ else 1; added_=0
                        with st.spinner("Adding issues..."):
                            for i,(_,ir) in enumerate(sr_.iterrows()):
                                ok=add_issue_to_sheet([f"ISS-{str(sn_+i).zfill(3)}",str(ir.get("site_name","")).strip(),str(ir.get("issue_type","Other")).strip(),str(ir.get("severity","Medium")).strip(),str(ir.get("description","")).strip(),str(ir.get("raised_by",st.session_state.get("user_email",""))).strip(),str(ir.get("raised_date",datetime.now().strftime("%d-%m-%Y"))).strip(),"","","Open","",datetime.now().strftime("%d-%m-%Y %H:%M"),""])
                                if ok: added_+=1
                        if added_>0: st.success(f"✅ {added_} issues added!"); st.session_state["analyzed_issues"]=[]; st.rerun()
                        else: st.error("Failed to add. Check sheet access.")
                else: st.info("Nothing selected.")

# ==========================================
# TAB 2: MASTER PROJECT ANALYTICS
# ==========================================
with tab_master:
    if master_df.empty:
        st.warning("No Master Project data found.")
    else:
        cs=safe_col(master_df,["STATE","State"]); cd=safe_col(master_df,["DISTRICT / CITY","DISTRICT","District"])
        cst=safe_col(master_df,["STATUS OF PROJECT","Status","STATUS"]); cte=safe_col(master_df,["Technical Person","TECHNICAL PERSON NAME","TECHNICAL PERSON"])
        csa=safe_col(master_df,["Sells Person","SALES PERSON NAME","SALES PERSON","Sales Person"]); cdi=safe_col(master_df,["Distributer","DISTRIBUTOR NANE","DISTRIBUTOR","Distributor"])
        con=safe_col(master_df,["VISIT ONGOING","Visit Ongoing"])
        st.subheader("Master Filters")
        mc1,mc2,mc3,mc4,mc5,mc6=st.columns(6); fm=master_df.copy()
        if cs: fst=mc1.selectbox("State",["All"]+clean_options(fm[cs]),key="t2_state"); fm=fm[fm[cs].astype(str)==fst] if fst!="All" else fm
        if cd: fdi=mc2.selectbox("District",["All"]+clean_options(fm[cd]),key="t2_dist"); fm=fm[fm[cd].astype(str)==fdi] if fdi!="All" else fm
        if cst: fsp=mc3.selectbox("Project Status",["All"]+clean_options(fm[cst]),key="t2_stat"); fm=fm[fm[cst].astype(str)==fsp] if fsp!="All" else fm
        if cte: fte=mc4.selectbox("Tech Person",["All"]+clean_options(fm[cte]),key="t2_tech"); fm=fm[fm[cte].astype(str)==fte] if fte!="All" else fm
        if csa: fsa=mc5.selectbox("Sales Person",["All"]+clean_options(fm[csa]),key="t2_sale"); fm=fm[fm[csa].astype(str)==fsa] if fsa!="All" else fm
        if cdi: fdr=mc6.selectbox("Distributor",["All"]+clean_options(fm[cdi]),key="t2_dist2"); fm=fm[fm[cdi].astype(str)==fdr] if fdr!="All" else fm
        ts_=set(); [ts_.update(fm[c].dropna().astype(str).tolist()) for c in [cte,csa] if c]
        k1,k2,k3,k4=st.columns(4)
        k1.metric("Total Projects",len(fm))
        k2.metric("Active (Ongoing)",len(fm[fm[con].astype(str).str.lower().isin(["yes","y","ongoing"])]) if con else 0)
        k3.metric("States Covered",fm[cs].nunique() if cs else 0)
        k4.metric("Tech/Sales Teams",len([x for x in ts_ if x.strip() and x.lower() not in ["nan","none",""]]))
        st.markdown("---")
        mc_a,mc_b=st.columns(2)
        with mc_a:
            with st.container(border=True):
                st.markdown("##### Projects by State")
                if cs: sc_=fm[cs].value_counts().reset_index(); sc_.columns=["State","Count"]; st.plotly_chart(px.bar(sc_,x="State",y="Count",color_discrete_sequence=["#14b8a6"]),use_container_width=True,key="c_t2s")
        with mc_b:
            with st.container(border=True):
                st.markdown("##### Project Status")
                if cst: ss_=fm[cst].value_counts().reset_index(); ss_.columns=["Status","Count"]; st.plotly_chart(px.pie(ss_,names="Status",values="Count",hole=0.4,color_discrete_sequence=["#6366f1","#14b8a6","#f59e0b","#f43f5e"]),use_container_width=True,key="c_t2p")
        st.subheader("Master Projects Directory")
        st.dataframe(fm.astype(str),use_container_width=True,hide_index=True)

# ==========================================
# TAB 3: EXECUTIVE DASHBOARD
# ==========================================
with tab_exec:
    ec1,ec2=st.columns([4,1])
    with ec1: st.markdown("### Executive Dashboard"); st.markdown("Multi-month associate performance tracking & field analytics")
    with ec2:
        sel_month=st.selectbox("Month",["All"]+clean_options(visits_df["Month"]),label_visibility="collapsed",key="t3_month") if not visits_df.empty else "All"
    if visits_df.empty:
        st.warning("No Visit Log data found.")
    else:
        edf=visits_df.copy()
        if sel_month!="All": edf=edf[edf["Month"]==sel_month]
        ae=safe_col(edf,["Associate ID","Associate","Technical Person"]); se=find_visit_site_col(edf)
        if not ae:
            st.error("Associate ID column not found.")
        else:
            sr=[]
            for assoc,grp in edf.groupby(ae):
                if pd.isna(assoc) or str(assoc).strip()=="": continue
                my=grp["Clean_Report_Mark"].isin(["YES","Y","TRUE"]); mn=grp["Clean_Report_Mark"].isin(["NO","N","FALSE"])
                sr.append({"Associate ID":assoc,"Floor Visit":int(grp["Num_Floors"].sum()),"Site Tower visit":int(grp[se].count() if se else len(grp)),"Report Mark (YES)":int(grp[my]["Num_Floors"].sum()),"Suggestion Visit (NO)":int(grp[mn]["Num_Floors"].sum()),"Report Pending":len(grp[grp["Status"]=="Pending"]),"Report sent to the client":int(grp[my]["Num_Floors"].sum()),"March Month(Pending)":0,"Report total with Pend":int(grp[my]["Num_Floors"].sum())})
            sdf=pd.DataFrame(sr)
            tf=sdf["Floor Visit"].sum() if not sdf.empty else 0; ts_v=sdf["Site Tower visit"].sum() if not sdf.empty else 0
            tse=sdf["Report sent to the client"].sum() if not sdf.empty else 0; tp=sdf["Report Pending"].sum() if not sdf.empty else 0
            k1,k2,k3,k4=st.columns(4)
            k1.metric("TOTAL FLOOR VISITS",tf); k2.metric("TOTAL SITE VISITS",ts_v); k3.metric("TOTAL REPORTS SENT",tse); k4.metric("TOTAL PENDING REPORTS",tp)
            st.markdown("---")
            if "exec_ctr" not in st.session_state: st.session_state["exec_ctr"]=0
            ctr=st.session_state["exec_ctr"]; sel_assoc=None
            ch_a,ch_b=st.columns(2)
            with ch_a:
                with st.container(border=True):
                    st.markdown("#### 📊 Reports Sent to Client"); st.caption("👆 Click bar to filter table")
                    if not sdf.empty:
                        s1=sdf.sort_values("Report sent to the client",ascending=True)
                        fl=px.bar(s1,x="Report sent to the client",y="Associate ID",orientation="h",text="Report sent to the client",color_discrete_sequence=["#3b82f6"])
                        fl.update_traces(textposition="outside"); fl.update_layout(xaxis_title="",yaxis_title="",showlegend=False,margin=dict(l=0,r=0,t=30,b=0))
                        ev1=st.plotly_chart(fl,use_container_width=True,key=f"c_t3r_{ctr}",on_select="rerun",selection_mode="points")
                        if ev1 and ev1.get("selection",{}).get("points"): pts=ev1["selection"]["points"]; sel_assoc=pts[0].get("y") if pts else None
            with ch_b:
                with st.container(border=True):
                    st.markdown("#### 🏢 Floor vs Site Breakdown"); st.caption("👆 Click bar to filter table")
                    if not sdf.empty:
                        dm=sdf.melt(id_vars="Associate ID",value_vars=["Floor Visit","Site Tower visit"],var_name="Visit Type",value_name="Count")
                        fr=px.bar(dm,x="Count",y="Associate ID",color="Visit Type",barmode="group",orientation="h",color_discrete_map={"Floor Visit":"#6366f1","Site Tower visit":"#10b981"})
                        fr.update_layout(xaxis_title="",yaxis_title="",legend_title="",margin=dict(l=0,r=0,t=30,b=0))
                        ev2=st.plotly_chart(fr,use_container_width=True,key=f"c_t3b_{ctr}",on_select="rerun",selection_mode="points")
                        if not sel_assoc and ev2 and ev2.get("selection",{}).get("points"): pts=ev2["selection"]["points"]; sel_assoc=pts[0].get("y") if pts else None
            st.markdown("#### 📋 Detailed Performance Breakdown")
            if not sdf.empty:
                tr=pd.DataFrame([{"Associate ID":"TEAM TOTALS","Floor Visit":tf,"Site Tower visit":ts_v,"Report Mark (YES)":sdf["Report Mark (YES)"].sum(),"Suggestion Visit (NO)":sdf["Suggestion Visit (NO)"].sum(),"Report Pending":tp,"Report sent to the client":tse,"March Month(Pending)":0,"Report total with Pend":sdf["Report total with Pend"].sum()}])
                ddf=pd.concat([sdf,tr],ignore_index=True)
                if sel_assoc:
                    fc1,fc2=st.columns([5,1])
                    fc1.info(f"🔎 Filtered: **{sel_assoc}**")
                    with fc2:
                        if st.button("✖ Clear",key="clr_exec",use_container_width=True): st.session_state["exec_ctr"]+=1; st.rerun()
                    show=ddf[ddf["Associate ID"]==sel_assoc]; show=ddf if show.empty else show
                else: show=ddf
                st.dataframe(show,use_container_width=True,hide_index=True)
                hc_="None"; hp_="None"; cg_="None"
                if len(sdf)>0:
                    im=sdf["Site Tower visit"].idxmax(); hc_=f"{sdf.loc[im,'Associate ID']} ({sdf.loc[im,'Site Tower visit']} Sites)"
                    jm=sdf["Floor Visit"].idxmax(); hp_=f"{sdf.loc[jm,'Associate ID']} ({sdf.loc[jm,'Floor Visit']} Floors)"
                    z0=sdf[sdf["Report sent to the client"]==0]; cg_=", ".join(z0["Associate ID"].astype(str).tolist())+" (0 Sent)" if not z0.empty else "All Associates Active"
                h1,h2,h3=st.columns(3)
                h1.markdown(f'<div class="highlight-card card-blue"><div class="card-title">🌎 Highest Coverage</div><div class="card-value">{hc_}</div></div>',unsafe_allow_html=True)
                h2.markdown(f'<div class="highlight-card card-green"><div class="card-title">🚀 Highest Productivity</div><div class="card-value">{hp_}</div></div>',unsafe_allow_html=True)
                h3.markdown(f'<div class="highlight-card card-red"><div class="card-title">⏳ Critical Gaps</div><div class="card-value">{cg_}</div></div>',unsafe_allow_html=True)
                st.markdown("---")
                pdf_c,pdf_d=st.columns([1,3])
                with pdf_c:
                    st.download_button("📄 Download PDF Report",data=build_executive_pdf(ddf,tf,ts_v,tse,tp,sel_month,hc_,hp_,cg_,sdf),file_name=f"Executive_{sel_month.replace(' ','_')}_{datetime.now().strftime('%d%m%Y')}.pdf",mime="application/pdf",key="dl_exec_pdf")
                with pdf_d:
                    st.caption("Print-ready PDF — KPIs, chart, full breakdown, highlights.")

# ==========================================
# TAB 4: SITE REPORT CARD
# ==========================================
with tab_site_card:
    st.markdown("### 🏢 Site Report Card")
    st.markdown("Select site + filters → card and KPIs update instantly.")
    if master_df.empty and visits_df.empty:
        st.warning("No data found.")
    else:
        msc=find_master_site_col(master_df); vsc=find_visit_site_col(visits_df)
        all_sites=sorted(list(set([x for x in (clean_options(master_df[msc]) if msc else [])+(clean_options(visits_df[vsc]) if vsc else []) if str(x).strip()])))
        if not all_sites:
            st.warning("No site names found.")
        else:
            sa1,sa2=st.columns([3,1])
            with sa1: sel_site=st.selectbox("Select Site Name",all_sites,key="sc_site")
            with sa2: st.write(""); st.write(""); show_all=st.checkbox("Show all columns",value=True,key="sc_all")
            sm=filter_site(master_df,msc,sel_site) if msc else pd.DataFrame()
            sv=filter_site(visits_df,vsc,sel_site) if vsc else pd.DataFrame()
            mr=sm.iloc[0] if not sm.empty else pd.Series(dtype="object")
            cp=msc; cst=safe_col(master_df,["STATE","State"]); cdi=safe_col(master_df,["DISTRICT / CITY","DISTRICT","District","CITY","City"])
            car=safe_col(master_df,["Area","AREA"]); csp=safe_col(master_df,["STATUS OF PROJECT","Status","STATUS"])
            cvo=safe_col(master_df,["VISIT ONGOING","Visit Ongoing"]); cte=safe_col(master_df,["Technical Person","TECHNICAL PERSON NAME","TECHNICAL PERSON"])
            csa=safe_col(master_df,["Sells Person","SALES PERSON NAME","SALES PERSON","Sales Person"]); cdr=safe_col(master_df,["Distributer","DISTRIBUTOR NANE","DISTRIBUTOR","Distributor"])
            asc=safe_col(sv,["Associate ID","Associate","Technical Person"]); dcs=safe_col(sv,["Date of Visit","Visit Date","Date"])
            cmc=safe_col(sv,["Comment","Remarks","Observation"]); twc=safe_col(sv,["Tower Name","Tower","Building"])
            flc=safe_col(sv,["FloorsVisited","Floors Visited","Floor Visited","Floor"])
            if not sv.empty:
                st.markdown("#### 🔽 Filter View")
                f1,f2,f3,f4=st.columns(4)
                sfm=f1.selectbox("Month",["All"]+(clean_options(sv["Month"]) if "Month" in sv.columns else []),key="sc_mon")
                sft=f2.selectbox("Tower",["All"]+(clean_options(sv[twc]) if twc else []),key="sc_tow")
                sfa=f3.selectbox("Associate",["All"]+(clean_options(sv[asc]) if asc else []),key="sc_asc")
                sfs=f4.selectbox("Status",["All"]+(clean_options(sv["Status"]) if "Status" in sv.columns else []),key="sc_sts")
                svf=sv.copy()
                if sfm!="All" and "Month" in svf.columns: svf=svf[svf["Month"]==sfm]
                if sft!="All" and twc: svf=svf[svf[twc].astype(str)==sft]
                if sfa!="All" and asc: svf=svf[svf[asc].astype(str)==sfa]
                if sfs!="All" and "Status" in svf.columns: svf=svf[svf["Status"]==sfs]
            else:
                sfm=sft=sfa=sfs="All"; svf=sv.copy()
            tvr=len(svf); tfl=int(svf["Num_Floors"].sum()) if not svf.empty and "Num_Floors" in svf.columns else 0
            sr_=len(svf[svf["Status"]=="Submitted"]) if not svf.empty and "Status" in svf.columns else 0
            pr_=len(svf[svf["Status"]=="Pending"]) if not svf.empty and "Status" in svf.columns else 0
            tna_=len(svf[svf["Status"]=="Technical (NA)"]) if not svf.empty and "Status" in svf.columns else 0
            tt_=svf[twc].nunique() if twc and not svf.empty else 0
            lvd=lvb=lvc="-"
            if not svf.empty:
                srt=svf.sort_values("Date Parsed",ascending=False) if "Date Parsed" in svf.columns else svf
                lr=srt.iloc[0]
                if dcs: lvd=lr.get(dcs,"-")
                if asc: lvb=lr.get(asc,"-")
                if cmc: lvc=lr.get(cmc,"-")
            fp=[x for x in [(sfm if sfm!="All" else None),(sft if sft!="All" else None),(sfa if sfa!="All" else None),(sfs if sfs!="All" else None)] if x]
            fl_="|".join(fp) if fp else "All Data"
            mrr=mr.copy(); mrr["Last Visit Date"]=lvd; mrr["Last Visit By"]=lvb
            mc1=[("Project / Site Name",cp),("State",cst),("District / City",cdi),("Area",car),("Project Status",csp),("Visit Ongoing",cvo)]
            mc2=[("Technical Person",cte),("Sales Person",csa),("Distributor",cdr),("Source Sheet","Source Sheet"),("Last Visit Date","Last Visit Date"),("Last Visit By","Last Visit By")]
            sdf_=pd.DataFrame([{"Site Name":sel_site,"Total Visit Records":tvr,"Total Floor Visits":tfl,"Submitted Reports":sr_,"Pending Reports":pr_,"Technical NA":tna_,"Total Towers":tt_,"Last Visit Date":lvd,"Last Visit By":lvb,"Filter Applied":fl_}])
            components.html(create_site_card_html(sel_site,mrr,mc1,mc2,tvr,tfl,sr_,pr_,tna_,tt_,lvd,lvb,lvc,filter_label=fl_),height=640,scrolling=True)
            st.markdown("### 📋 VisitLog Data")
            if svf.empty:
                st.info(f"No records match filters: **{fl_}**")
            else:
                pc=[c for c in ["Source Sheet","Visit ID",vsc,twc,flc,asc,dcs,"Is Report Visit?","Status","Report Submitted Date",cmc,"CreatedAt"] if c and c in svf.columns]
                pc=list(dict.fromkeys(pc)); vdf=svf.copy() if show_all else svf[pc].copy()
                cc_a,cc_b=st.columns(2)
                with cc_a:
                    with st.container(border=True):
                        st.markdown("##### Visits by Month")
                        if "Month" in svf.columns:
                            mc_=svf["Month"].value_counts().reset_index(); mc_.columns=["Month","Visits"]
                            st.plotly_chart(px.bar(mc_,x="Month",y="Visits",color_discrete_sequence=["#6366f1"]),use_container_width=True,key="c_sc_m")
                with cc_b:
                    with st.container(border=True):
                        st.markdown("##### Status Breakdown")
                        if "Status" in svf.columns:
                            sc_=svf["Status"].value_counts().reset_index(); sc_.columns=["Status","Count"]
                            st.plotly_chart(px.pie(sc_,names="Status",values="Count",hole=0.4),use_container_width=True,key="c_sc_s")
                st.dataframe(vdf.astype(str),use_container_width=True,hide_index=True)
                st.markdown("### 📌 Full MasterProject Data")
                mdd=sm.copy() if not sm.empty else pd.DataFrame()
                if sm.empty: st.warning("Site not found in MasterProject.")
                else: st.dataframe(mdd.astype(str),use_container_width=True,hide_index=True)
                sfn=sel_site.replace("/","_").replace("\\","_").replace(" ","_")
                st.markdown("---"); st.caption("📥 Downloads = **complete visit history** for this site (all towers/months), not filtered view.")
                all_cols=[c for c in ["Source Sheet","Visit ID",vsc,twc,flc,asc,dcs,"Is Report Visit?","Status","Report Submitted Date",cmc] if c and c in sv.columns]; all_cols=list(dict.fromkeys(all_cols))
                dlvdf=sv[all_cols].copy() if all_cols else sv.copy()
                d1,d2,d3=st.columns(3)
                d1.download_button("⬇️ Download Excel",data=create_excel_compatible_report(sel_site,mdd,dlvdf,sdf_,lvc),file_name=f"{sfn}_Report.xls",mime="application/vnd.ms-excel",key="dl_sc_xl")
                d2.download_button("🖨️ Download Print Report",data=create_print_html_report(sel_site,mrr,mc1,mc2,sdf_,dlvdf,lvc),file_name=f"{sfn}_Print.html",mime="text/html",key="dl_sc_html")
                d3.download_button("📄 Download CSV",data=dlvdf.to_csv(index=False).encode("utf-8"),file_name=f"{sfn}_VisitLog.csv",mime="text/csv",key="dl_sc_csv")

# ==========================================
# TAB 5: SITE ISSUES
# ==========================================
with tab_issues:
    st.markdown("### 🚨 Site Issue Tracker")
    if st.button("🔄 Refresh Issues",key="refresh_issues"): st.cache_data.clear(); st.rerun()
    # issues_df already loaded globally
    def _cs(df,s): return 0 if df.empty or "Status" not in df.columns else len(df[df["Status"]==s])
    def _cv(df,s): return 0 if df.empty or "Severity" not in df.columns else len(df[df["Severity"].str.contains(s,case=False,na=False)])
    ik1,ik2,ik3,ik4,ik5=st.columns(5)
    ik1.metric("🔴 Open",_cs(issues_df,"Open")); ik2.metric("🟡 In Progress",_cs(issues_df,"In Progress"))
    ik3.metric("🟢 Resolved",_cs(issues_df,"Resolved")); ik4.metric("⚫ Closed",_cs(issues_df,"Closed"))
    ik5.metric("🔥 High Severity",_cv(issues_df,"High"))
    st.markdown("---")
    sl1,sl2,sl3=st.tabs(["📋 Issue Log","➕ Raise New Issue","📊 Issue Analytics"])
    _ms=find_master_site_col(master_df); _vs=find_visit_site_col(visits_df)
    asi=sorted(list(set([x for x in (clean_options(master_df[_ms]) if _ms and not master_df.empty else [])+(clean_options(visits_df[_vs]) if _vs and not visits_df.empty else []) if str(x).strip()])))
    _ac=safe_col(visits_df,["Associate ID","Associate","Technical Person"])
    al=clean_options(visits_df[_ac]) if _ac and not visits_df.empty else []
    with sl1:
        if issues_df.empty:
            st.info("No issues yet. Use 'Raise New Issue' tab.")
        else:
            fi1,fi2,fi3,fi4=st.columns(4)
            fis=fi1.selectbox("Site Name",["All"]+(clean_options(issues_df["Site Name"]) if "Site Name" in issues_df.columns else []),key="if_site")
            fist=fi2.selectbox("Status",["All"]+STATUS_OPTIONS_ISSUE,key="if_stat")
            fisv=fi3.selectbox("Severity",["All"]+(clean_options(issues_df["Severity"]) if "Severity" in issues_df.columns else []),key="if_sev")
            fity=fi4.selectbox("Issue Type",["All"]+ISSUE_TYPES,key="if_type")
            fi=issues_df.copy()
            if fis!="All" and "Site Name" in fi.columns: fi=fi[fi["Site Name"]==fis]
            if fist!="All" and "Status" in fi.columns: fi=fi[fi["Status"]==fist]
            if fisv!="All" and "Severity" in fi.columns: fi=fi[fi["Severity"]==fisv]
            if fity!="All" and "Issue Type" in fi.columns: fi=fi[fi["Issue Type"]==fity]
            lc=[c for c in ["Issue ID","Site Name","Issue Type","Severity","Description","Status","Raised By","Raised Date","Assigned To","Target Date","Resolution Notes","Created At","Updated At"] if c in fi.columns]
            st.dataframe(fi[lc].astype(str),use_container_width=True,hide_index=True)
            st.markdown(f"*Showing {len(fi)} of {len(issues_df)} total issues*")
            dl_a,dl_b=st.columns([1,3])
            with dl_a: st.download_button("📥 Download Excel (Color-Coded)",data=build_issues_excel(fi),file_name=f"Site_Issues_{datetime.now().strftime('%d%m%Y')}.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",key="dl_issues_xl")
            with dl_b: st.caption("🟢 Resolved  🟡 In Progress  🔴 Open  ⚪ Closed — rows color-coded by status.")
            st.markdown("---")
            with st.expander("✏️ Update Issue Status",expanded=False):
                if "Issue ID" not in issues_df.columns:
                    st.info("No issues to update.")
                else:
                    upd=[x for x in (issues_df[issues_df["Status"].isin(["Open","In Progress"])]["Issue ID"].dropna().tolist() if "Status" in issues_df.columns else issues_df["Issue ID"].dropna().tolist())]
                    if not upd: st.success("✅ All issues resolved or closed.")
                    else:
                        u1,u2=st.columns(2)
                        with u1:
                            sid=st.selectbox("Select Issue ID",upd,key="upd_id")
                            if sid and not issues_df.empty:
                                m=issues_df[issues_df["Issue ID"]==sid]
                                if not m.empty: dv=m.iloc[0].get("Description","-"); st.caption(f"📝 {str(dv)[:120]}{'...' if len(str(dv))>120 else ''}")
                        with u2: ns=st.selectbox("New Status",STATUS_OPTIONS_ISSUE,key="upd_stat")
                        rn=st.text_area("Resolution Notes",placeholder="What was done to resolve this?",key="upd_note",height=80)
                        if st.button("✅ Update Issue",key="btn_upd",use_container_width=True):
                            with st.spinner("Updating..."): ok=update_issue_in_sheet(sid,ns,rn)
                            if ok: st.success(f"✅ {sid} → '{ns}'"); st.rerun()
                            else: st.error("Update failed. Check sheet access.")
    with sl2:
        st.markdown("#### Raise a New Site Issue")
        fi_=load_issues(); nid=generate_issue_id(fi_); st.info(f"🆔 New Issue ID: **{nid}**")
        with st.form("raise_form",clear_on_submit=True):
            ra1,ra2=st.columns(2)
            with ra1:
                isite=st.selectbox("Site Name *",asi if asi else ["— No sites —"],key="ni_site")
                itype=st.selectbox("Issue Type *",ISSUE_TYPES,key="ni_type")
                isev=st.selectbox("Severity *",["🔴 High","🟡 Medium","🟢 Low"],key="ni_sev")
                irb=st.text_input("Raised By *",value=st.session_state.get("user_email",""),key="ni_rb")
            with ra2:
                ird=st.date_input("Raised Date *",value=datetime.today(),key="ni_rd")
                iat=st.selectbox("Assign To",["— Unassigned —"]+al,key="ni_at") if al else st.text_input("Assign To",placeholder="Name of person",key="ni_at_txt")
                itd=st.date_input("Target Resolution Date",value=datetime.today(),key="ni_td")
            idesc=st.text_area("Issue Description *",placeholder="Describe clearly — location, condition, NBC/Huliot non-compliance...",height=130,key="ni_desc")
            if st.form_submit_button("🚨 Raise Issue",use_container_width=True):
                if not idesc.strip(): st.error("Description is required.")
                elif not irb.strip(): st.error("Raised By is required.")
                else:
                    sc_=isev.split(" ",1)[-1].strip() if " " in isev else isev
                    iat_=(iat if isinstance(iat,str) else "").replace("— Unassigned —","").strip()
                    with st.spinner("Saving..."): ok=add_issue_to_sheet([nid,str(isite),str(itype),sc_,idesc.strip(),irb.strip(),str(ird),iat_,str(itd),"Open","",datetime.now().strftime("%d-%m-%Y %H:%M"),""])
                    if ok: st.success(f"✅ Issue **{nid}** raised! Refresh to see in log.")
                    else: st.error("Failed to save.")
    with sl3:
        if issues_df.empty:
            st.info("No issues yet.")
        else:
            sc_a,sc_b=st.columns(2)
            with sc_a:
                with st.container(border=True):
                    st.markdown("##### Issues by Status")
                    if "Status" in issues_df.columns:
                        ss=issues_df["Status"].value_counts().reset_index(); ss.columns=["Status","Count"]
                        st.plotly_chart(px.pie(ss,names="Status",values="Count",hole=0.4,color_discrete_map={"Open":"#f43f5e","In Progress":"#f59e0b","Resolved":"#22c55e","Closed":"#94a3b8"}),use_container_width=True,key="c_is_s")
            with sc_b:
                with st.container(border=True):
                    st.markdown("##### Issues by Severity")
                    if "Severity" in issues_df.columns:
                        sv_=issues_df["Severity"].value_counts().reset_index(); sv_.columns=["Severity","Count"]
                        fg=px.bar(sv_,x="Severity",y="Count",color="Severity",color_discrete_map={"High":"#f43f5e","Medium":"#f59e0b","Low":"#22c55e"}); fg.update_layout(showlegend=False)
                        st.plotly_chart(fg,use_container_width=True,key="c_is_v")
            sc_c,sc_d=st.columns(2)
            with sc_c:
                with st.container(border=True):
                    st.markdown("##### Top Sites — Open Issues")
                    if "Site Name" in issues_df.columns and "Status" in issues_df.columns:
                        od=issues_df[issues_df["Status"].isin(["Open","In Progress"])]
                        if not od.empty:
                            ts_=od["Site Name"].value_counts().nlargest(8).reset_index(); ts_.columns=["Site","Count"]
                            fg=px.bar(ts_,x="Count",y="Site",orientation="h",color_discrete_sequence=["#f43f5e"]); fg.update_layout(yaxis=dict(autorange="reversed"))
                            st.plotly_chart(fg,use_container_width=True,key="c_is_t")
                        else: st.success("✅ No open issues!")
            with sc_d:
                with st.container(border=True):
                    st.markdown("##### Issues by Type")
                    if "Issue Type" in issues_df.columns:
                        it_=issues_df["Issue Type"].value_counts().reset_index(); it_.columns=["Type","Count"]
                        st.plotly_chart(px.pie(it_,names="Type",values="Count",hole=0.4,color_discrete_sequence=px.colors.qualitative.Pastel),use_container_width=True,key="c_is_ty")
            if "Assigned To" in issues_df.columns and "Status" in issues_df.columns:
                oa=issues_df[issues_df["Status"].isin(["Open","In Progress"])&(issues_df["Assigned To"].str.strip()!="")]
                if not oa.empty:
                    with st.container(border=True):
                        st.markdown("##### Open Issues by Assignee")
                        ac_=oa["Assigned To"].value_counts().reset_index(); ac_.columns=["Assignee","Open Issues"]
                        fg=px.bar(ac_,x="Open Issues",y="Assignee",orientation="h",color_discrete_sequence=["#6366f1"],text="Open Issues"); fg.update_traces(textposition="outside"); fg.update_layout(yaxis=dict(autorange="reversed"),margin=dict(l=0,r=60,t=30,b=0))
                        st.plotly_chart(fg,use_container_width=True,key="c_is_a")

# ==========================================
# TAB 6: SITE MAP — Leaflet.js
# Improvements over old Plotly version:
# • Scroll-wheel zoom + full pan (native browser)
# • Google Maps-quality CartoDB dark tiles (free, no key)
# • 3-level precision: State → City → Neighborhood/Area
# • MarkerCluster — clean for dense cities
# • City dropdown zooms map instantly
# • Click marker → rich popup (all project details)
# • Tech-person filter highlights their sites
# • Open Issues count per site (from SiteIssues sheet)
# • Color-coded by Status + match-level badge
# ==========================================
with tab_map:
    st.markdown("### 🗺️ West Zone Project Map")
    st.markdown("All MasterProject sites geocoded to **neighborhood level** where project name or city matches a known area. Scroll to zoom. Click marker for details.")

    if master_df.empty:
        st.warning("No MasterProject data found.")
    else:
        mc_s  = find_master_site_col(master_df)
        mc_st = safe_col(master_df, ["STATE","State"])
        mc_di = safe_col(master_df, ["DISTRICT / CITY","DISTRICT","District","CITY","City"])
        mc_sp = safe_col(master_df, ["STATUS OF PROJECT","Status","STATUS"])
        mc_te = safe_col(master_df, ["Technical Person","TECHNICAL PERSON NAME","TECHNICAL PERSON"])
        mc_sa = safe_col(master_df, ["Sells Person","SALES PERSON NAME","SALES PERSON","Sales Person"])
        mc_on = safe_col(master_df, ["VISIT ONGOING","Visit Ongoing"])
        mc_di2= safe_col(master_df, ["Distributer","DISTRIBUTOR NANE","DISTRIBUTOR","Distributor"])

        if not mc_s:
            st.warning("Project/Site Name column not found in MasterProject.")
        else:
            # ── Filters ──────────────────────────────────────────
            mf1, mf2, mf3, mf4 = st.columns(4)
            with mf1:
                map_states = ["All"] + (clean_options(master_df[mc_st]) if mc_st else [])
                f_mst = st.selectbox("State", map_states, key="map_state")
            with mf2:
                _city_opts = ["All"] + (clean_options(master_df[mc_di]) if mc_di else [])
                f_mci = st.selectbox("City / District", _city_opts, key="map_city")
            with mf3:
                _stat_opts = ["All"] + (clean_options(master_df[mc_sp]) if mc_sp else [])
                f_msp = st.selectbox("Project Status", _stat_opts, key="map_status")
            with mf4:
                _tech_opts = ["All"] + (clean_options(master_df[mc_te]) if mc_te else [])
                f_mte = st.selectbox("Technical Person", _tech_opts, key="map_tech")

            fmap = master_df.copy()
            if mc_st and f_mst != "All": fmap = fmap[fmap[mc_st].astype(str) == f_mst]
            if mc_di and f_mci != "All": fmap = fmap[fmap[mc_di].astype(str) == f_mci]
            if mc_sp and f_msp != "All": fmap = fmap[fmap[mc_sp].astype(str) == f_msp]
            if mc_te and f_mte != "All": fmap = fmap[fmap[mc_te].astype(str) == f_mte]

            # ── Build open-issue counts per site ──────────────────
            issue_counts = {}
            if not issues_df.empty and "Site Name" in issues_df.columns and "Status" in issues_df.columns:
                oi = issues_df[issues_df["Status"].isin(["Open","In Progress"])]
                issue_counts = oi["Site Name"].value_counts().to_dict()

            # ── Geocode every row ─────────────────────────────────
            sites_json = []
            unmatched  = []
            for _, row in fmap.iterrows():
                proj  = str(row.get(mc_s,  "")).strip()
                state = str(row.get(mc_st, "")).strip() if mc_st else ""
                city  = str(row.get(mc_di, "")).strip() if mc_di else ""
                status= str(row.get(mc_sp, "")).strip() if mc_sp else "Unknown"
                tech  = str(row.get(mc_te, "")).strip() if mc_te else ""
                sales = str(row.get(mc_sa, "")).strip() if mc_sa else ""
                dist  = str(row.get(mc_di2,"")).strip() if mc_di2 else ""
                ongoing=str(row.get(mc_on, "")).strip() if mc_on else ""
                if not proj: continue
                lat, lon, level = geocode_site(proj, city, state)
                if lat is None:
                    unmatched.append(proj); continue
                # tiny jitter so overlapping city-level markers don't stack
                import hashlib
                h = int(hashlib.md5(proj.encode()).hexdigest(), 16)
                if level != "area":
                    lat += ((h % 17) - 8) * 0.008
                    lon += ((h % 13) - 6) * 0.008
                open_issues = issue_counts.get(proj, 0)
                sites_json.append({
                    "name": proj, "state": state or "-", "city": city or "-",
                    "status": status or "Unknown", "tech": tech or "-",
                    "sales": sales or "-", "dist": dist or "-",
                    "ongoing": ongoing or "-", "level": level,
                    "open_issues": open_issues,
                    "lat": round(lat, 6), "lon": round(lon, 6),
                    "highlight": (f_mte != "All" and tech == f_mte),
                })

            # ── KPI strip ─────────────────────────────────────────
            mk1,mk2,mk3,mk4 = st.columns(4)
            mk1.metric("Sites Plotted",     len(sites_json))
            mk2.metric("States Covered",    len(set(s["state"] for s in sites_json)))
            mk3.metric("Area-level precision", sum(1 for s in sites_json if s["level"]=="area"))
            mk4.metric("Unmatched", len(unmatched))

            if unmatched:
                with st.expander(f"⚠️ {len(unmatched)} sites could not be plotted"):
                    st.write(", ".join(unmatched[:60]) + (" ..." if len(unmatched)>60 else ""))
                    st.caption("Tip: check spelling of District/City column, or add a known area name to the project name.")

            st.markdown("---")

            if not sites_json:
                st.info("No sites to plot with current filters.")
            else:
                # ── Determine map centre & zoom from filters ───────
                if f_mci != "All" and f_mci.lower() in CITY_COORDS:
                    map_center = list(CITY_COORDS[f_mci.lower()])
                    map_zoom   = 12
                elif f_mst != "All" and f_mst.lower() in STATE_COORDS:
                    map_center = list(STATE_COORDS[f_mst.lower()])
                    map_zoom   = 8
                else:
                    # centroid of all plotted sites
                    map_center = [
                        round(sum(s["lat"] for s in sites_json)/len(sites_json),4),
                        round(sum(s["lon"] for s in sites_json)/len(sites_json),4)
                    ]
                    map_zoom = 6 if len(sites_json) > 30 else 7

                sites_js = json.dumps(sites_json, ensure_ascii=False)

                # ── Leaflet HTML ──────────────────────────────────
                leaflet_html = f"""<!DOCTYPE html>
<html>
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css"/>
<link rel="stylesheet" href="https://unpkg.com/leaflet.markercluster@1.5.3/dist/MarkerCluster.css"/>
<link rel="stylesheet" href="https://unpkg.com/leaflet.markercluster@1.5.3/dist/MarkerCluster.Default.css"/>
<script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
<script src="https://unpkg.com/leaflet.markercluster@1.5.3/dist/leaflet.markercluster.js"></script>
<style>
  *{{margin:0;padding:0;box-sizing:border-box;}}
  body{{background:#0E1B2E;font-family:Inter,-apple-system,sans-serif;}}
  #map{{width:100%;height:600px;}}
  .popup-card{{min-width:220px;font-family:Inter,-apple-system,sans-serif;font-size:12px;}}
  .popup-title{{font-size:14px;font-weight:700;color:#0E1B2E;margin-bottom:8px;border-bottom:2px solid #38BDF8;padding-bottom:5px;}}
  .popup-row{{display:flex;justify-content:space-between;margin-bottom:4px;}}
  .popup-label{{color:#6b7280;font-weight:600;margin-right:8px;white-space:nowrap;}}
  .popup-val{{color:#111827;text-align:right;font-weight:500;}}
  .popup-badge{{display:inline-block;padding:2px 8px;border-radius:10px;font-size:10px;font-weight:700;margin-top:5px;}}
  .popup-issues{{color:#B91C1C;background:#FEE2E2;}}
  .popup-level{{color:#1D4ED8;background:#EFF6FF;margin-left:4px;}}
  .leaflet-popup-content-wrapper{{border-radius:10px;box-shadow:0 8px 24px rgba(0,0,0,0.25);}}
  .legend{{background:rgba(14,27,46,0.90);backdrop-filter:blur(8px);padding:10px 14px;border-radius:10px;border:1px solid rgba(56,189,248,0.3);font-size:11px;font-family:Inter,sans-serif;color:#CBD5E1;line-height:1.8;}}
  .legend-title{{font-weight:700;font-size:12px;color:#38BDF8;margin-bottom:6px;}}
  .ldot{{display:inline-block;width:11px;height:11px;border-radius:50%;margin-right:5px;border:1.5px solid rgba(255,255,255,0.4);vertical-align:middle;}}
</style>
</head>
<body>
<div id="map"></div>
<script>
var SITES = {sites_js};
var CENTER = [{map_center[0]}, {map_center[1]}];
var ZOOM = {map_zoom};

var map = L.map('map', {{
  center: CENTER,
  zoom: ZOOM,
  scrollWheelZoom: true,
  zoomControl: true,
  preferCanvas: true
}});

// CartoDB Dark Matter tiles — Google Maps quality, free, no API key
L.tileLayer('https://{{s}}.basemaps.cartocdn.com/dark_all/{{z}}/{{x}}/{{y}}{{r}}.png', {{
  attribution: '&copy; <a href="https://www.openstreetmap.org/copyright">OSM</a> &copy; <a href="https://carto.com">CartoDB</a>',
  subdomains: 'abcd',
  maxZoom: 19
}}).addTo(map);

var STATUS_COLORS = {{
  'Completed': '#22c55e', 'Complete': '#22c55e', 'Done': '#22c55e',
  'Ongoing': '#38BDF8', 'In Progress': '#38BDF8', 'Active': '#38BDF8',
  'Pending': '#f59e0b', 'On Hold': '#f59e0b', 'Hold': '#f59e0b',
  'Cancelled': '#ef4444', 'Canceled': '#ef4444', 'Stopped': '#ef4444',
  'Unknown': '#94a3b8'
}};

function getColor(status) {{
  for (var k in STATUS_COLORS) {{
    if (status && status.toLowerCase() === k.toLowerCase()) return STATUS_COLORS[k];
  }}
  return '#94a3b8';
}}

function makeIcon(color, highlight, hasIssues) {{
  var size = highlight ? 18 : 13;
  var border = highlight ? '3px solid #FBBF24' : (hasIssues ? '2px solid #EF4444' : '2px solid rgba(255,255,255,0.5)');
  var glow = highlight ? 'box-shadow:0 0 10px #FBBF24,0 0 20px #FBBF24;' : (hasIssues ? 'box-shadow:0 0 8px #EF4444;' : '');
  var html = '<div style="width:'+size+'px;height:'+size+'px;border-radius:50%;background:'+color+';border:'+border+';'+glow+'"></div>';
  return L.divIcon({{html: html, className: '', iconSize: [size, size], iconAnchor: [Math.floor(size/2), Math.floor(size/2)]}});
}}

var clusterGroup = L.markerClusterGroup({{
  maxClusterRadius: 40,
  iconCreateFunction: function(cluster) {{
    var n = cluster.getChildCount();
    var sz = n > 20 ? 38 : n > 8 ? 32 : 26;
    return L.divIcon({{
      html: '<div style="width:'+sz+'px;height:'+sz+'px;border-radius:50%;background:rgba(56,189,248,0.85);border:2px solid white;display:flex;align-items:center;justify-content:center;font-size:'+Math.min(sz/2.2,14)+'px;font-weight:700;color:#0E1B2E;box-shadow:0 2px 8px rgba(0,0,0,0.4);">'+n+'</div>',
      className: '', iconSize: [sz, sz], iconAnchor: [sz/2, sz/2]
    }});
  }}
}});

SITES.forEach(function(s) {{
  var color = getColor(s.status);
  var icon  = makeIcon(color, s.highlight, s.open_issues > 0);
  var issuesBadge = s.open_issues > 0 ? '<span class="popup-badge popup-issues">🔴 '+s.open_issues+' open issue'+(s.open_issues>1?'s':'')+'</span>' : '<span style="color:#15803D;font-size:10px;font-weight:600;">✅ No open issues</span>';
  var levelBadge  = '<span class="popup-badge popup-level">'+(s.level==='area'?'📍 Area':'s.level'==='city'?'🏙 City':'🗺 State')+'</span>';
  var levelLabel  = s.level === 'area' ? '📍 Area-level' : s.level === 'city' ? '🏙 City-level' : '🗺 State-level';
  var popupHtml =
    '<div class="popup-card">'+
    '<div class="popup-title">'+s.name+'</div>'+
    '<div class="popup-row"><span class="popup-label">State</span><span class="popup-val">'+s.state+'</span></div>'+
    '<div class="popup-row"><span class="popup-label">City / District</span><span class="popup-val">'+s.city+'</span></div>'+
    '<div class="popup-row"><span class="popup-label">Status</span><span class="popup-val" style="color:'+color+';font-weight:700;">'+s.status+'</span></div>'+
    '<div class="popup-row"><span class="popup-label">Ongoing</span><span class="popup-val">'+s.ongoing+'</span></div>'+
    '<div class="popup-row"><span class="popup-label">Tech Person</span><span class="popup-val">'+s.tech+'</span></div>'+
    '<div class="popup-row"><span class="popup-label">Sales Person</span><span class="popup-val">'+s.sales+'</span></div>'+
    '<div class="popup-row"><span class="popup-label">Distributor</span><span class="popup-val">'+s.dist+'</span></div>'+
    '<div class="popup-row"><span class="popup-label">Precision</span><span class="popup-val" style="color:#1D4ED8;font-size:10px;">'+levelLabel+'</span></div>'+
    '<div style="margin-top:6px;">'+issuesBadge+'</div>'+
    '</div>';
  var marker = L.marker([s.lat, s.lon], {{icon: icon}});
  marker.bindPopup(popupHtml, {{maxWidth: 280, minWidth: 220}});
  clusterGroup.addLayer(marker);
}});

map.addLayer(clusterGroup);

// Legend
var legend = L.control({{position: 'bottomright'}});
legend.onAdd = function(map) {{
  var div = L.DomUtil.create('div', 'legend');
  div.innerHTML = '<div class="legend-title">Project Status</div>'+
    '<div><span class="ldot" style="background:#38BDF8;"></span>Ongoing / Active</div>'+
    '<div><span class="ldot" style="background:#22c55e;"></span>Completed / Done</div>'+
    '<div><span class="ldot" style="background:#f59e0b;"></span>Pending / On Hold</div>'+
    '<div><span class="ldot" style="background:#ef4444;"></span>Cancelled</div>'+
    '<div><span class="ldot" style="background:#94a3b8;"></span>Unknown</div>'+
    '<hr style="border-color:rgba(56,189,248,0.2);margin:6px 0;">'+
    '<div><span class="ldot" style="background:#FBBF24;border-color:#FBBF24;"></span>Selected Tech Person</div>'+
    '<div style="color:#EF4444;font-size:10px;margin-top:2px;">🔴 border = has open issues</div>'+
    '<hr style="border-color:rgba(56,189,248,0.2);margin:6px 0;">'+
    '<div style="font-size:10px;color:#94a3b8;">📍 Area-level | 🏙 City | 🗺 State</div>';
  return div;
}};
legend.addTo(map);
</script>
</body>
</html>"""

                with st.container(border=True):
                    components.html(leaflet_html, height=620, scrolling=False)

                st.caption(
                    "🟦 Ongoing  🟩 Completed  🟧 Pending  🟥 Cancelled  ⚪ Unknown  |  "
                    "🌟 Gold border = filtered Tech Person  🔴 Red border = has open issues  |  "
                    "📍 Area-level (most precise) > 🏙 City > 🗺 State  |  "
                    "Numbers on clusters = projects in that area — click to expand."
                )

                # ── Detail table ──────────────────────────────────
                st.markdown("---")
                st.subheader("Plotted Sites — Detail Table")
                tdf = pd.DataFrame(sites_json)[["name","state","city","status","tech","sales","ongoing","level","open_issues"]]
                tdf.columns = ["Project","State","City/District","Status","Tech Person","Sales Person","Ongoing","Precision","Open Issues"]
                st.dataframe(tdf, use_container_width=True, hide_index=True)

                # ── State summary bar chart ────────────────────────
                st.markdown("---")
                with st.container(border=True):
                    st.markdown("##### Projects by State")
                    sc_ = tdf["State"].value_counts().reset_index()
                    sc_.columns = ["State","Projects"]
                    st.plotly_chart(px.bar(sc_, x="State", y="Projects", color_discrete_sequence=["#38BDF8"]), use_container_width=True, key="map_state_bar")
